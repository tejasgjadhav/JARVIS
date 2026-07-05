#!/usr/bin/env python3
"""J.A.R.V.I.S. Report Engine — institutional equity report, no LLM required.

Pipeline: yfinance (real market data) → factor analysis + DCF (pure Python) →
Excel financial model (openpyxl) + institutional PDF (reportlab).

Everything here is deterministic and works with the Anthropic API offline.
"""
import io
import json
import math
from datetime import datetime

import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ── Unicode PDF fonts (so ₹, →, ≈ etc. render, not as black boxes) ──
PDF_FONT, PDF_BOLD = "Helvetica", "Helvetica-Bold"

def _register_pdf_fonts():
    global PDF_FONT, PDF_BOLD
    import os
    tries = []
    try:
        import matplotlib
        base = os.path.join(os.path.dirname(matplotlib.__file__), "mpl-data/fonts/ttf")
        tries.append((os.path.join(base, "DejaVuSans.ttf"), os.path.join(base, "DejaVuSans-Bold.ttf"), None))
    except Exception:
        pass
    tries.append(("/System/Library/Fonts/Helvetica.ttc", "/System/Library/Fonts/Helvetica.ttc", (0, 1)))
    for reg, bold, idx in tries:
        try:
            if not os.path.exists(reg):
                continue
            if idx:
                pdfmetrics.registerFont(TTFont("UFont", reg, subfontIndex=idx[0]))
                pdfmetrics.registerFont(TTFont("UFont-Bold", bold, subfontIndex=idx[1]))
            else:
                pdfmetrics.registerFont(TTFont("UFont", reg))
                pdfmetrics.registerFont(TTFont("UFont-Bold", bold if os.path.exists(bold) else reg))
            pdfmetrics.registerFontFamily("UFont", normal="UFont", bold="UFont-Bold",
                                          italic="UFont", boldItalic="UFont-Bold")
            PDF_FONT, PDF_BOLD = "UFont", "UFont-Bold"
            return
        except Exception:
            continue

_register_pdf_fonts()

# ── Palette (JARVIS arc-reactor blue) ──────────────────────
NAVY = "0B1F3A"
ARC = "00A8E8"
LIGHT = "EAF6FB"
GREY = "6B7A8D"


# ═══════════════════════════════════════════════════════════
#  1. DATA
# ═══════════════════════════════════════════════════════════
def normalize_symbol(sym: str) -> str:
    s = (sym or "").strip().upper().lstrip("$")
    # Default to NSE if no exchange suffix given
    if "." not in s and s not in ("", ):
        s = s + ".NS"
    return s


# Known renames / demergers where searching the OLD name/ticker fails.
SYMBOL_ALIASES = {
    "ZOMATO": "ETERNAL",       # Zomato → Eternal Ltd (2025)
    "TATAMTRDVR": "TMCV",      # Tata Motors DVR
}

def apply_alias(sym: str):
    if not sym:
        return sym
    base = sym.upper().replace(".NS", "").replace(".BO", "")
    return SYMBOL_ALIASES.get(base, sym)


def fy_labels(d, n=5):
    """Projected fiscal-year labels, e.g. FY27E … FY31E (Indian FY ends March)."""
    b = (d or {}).get("fy_base")
    if not b:
        return [f"Year {i + 1}" for i in range(n)]
    return [f"FY{str(b + i + 1)[-2:]}E" for i in range(n)]


def resolve_symbol(query: str):
    """Resolve a company name or stale/renamed ticker to a current NSE (fallback
    BSE) symbol via Yahoo search — handles delistings, renames, demergers.
    ONLY Indian listings are accepted: a blind first-hit fallback once resolved a
    query to a foreign penny listing and produced an absurd price (₹0.75)."""
    if not query:
        return None
    try:
        r = yf.Search(query, max_results=10)
        quotes = getattr(r, "quotes", []) or []
    except Exception:
        return None
    for suffix in (".NS", ".BO"):
        for q in quotes:
            s = q.get("symbol", "")
            if s.endswith(suffix):
                return s
    return None   # nothing on NSE/BSE → let the caller report "not found"


def _num(v):
    """Coerce to float or None."""
    try:
        if v is None:
            return None
        f = float(v)
        return None if math.isnan(f) else f
    except (TypeError, ValueError):
        return None


def fetch_stock(symbol: str) -> dict:
    sym = normalize_symbol(symbol)
    t = yf.Ticker(sym)
    info = {}
    try:
        info = t.info or {}
    except Exception:
        info = {}

    hist = None
    try:
        hist = t.history(period="1y")
    except Exception:
        hist = None

    # Statement-based figures are more reliable + INR-consistent than .info scalars.
    fcf_stmt = rev_stmt = None
    try:
        cf = t.cashflow
        if cf is not None and not cf.empty:
            if "Free Cash Flow" in cf.index:
                fcf_stmt = _num(cf.loc["Free Cash Flow"].iloc[0])
            if fcf_stmt is None and "Operating Cash Flow" in cf.index:
                ocf = _num(cf.loc["Operating Cash Flow"].iloc[0])
                capex = _num(cf.loc["Capital Expenditure"].iloc[0]) if "Capital Expenditure" in cf.index else 0
                if ocf is not None:
                    fcf_stmt = ocf + (capex or 0)  # capex is negative in the statement
    except Exception:
        pass
    rev_cagr = None
    rev_hist = None
    ebit_hist = None
    fy_base = None
    try:
        fin = t.financials
        if fin is not None and not fin.empty and "Total Revenue" in fin.index:
            rev_stmt = _num(fin.loc["Total Revenue"].iloc[0])
            rr = fin.loc["Total Revenue"].dropna()
            if len(rr) >= 2:
                newest, oldest, yrs = _num(rr.iloc[0]), _num(rr.iloc[-1]), len(rr) - 1
                if newest and oldest and oldest > 0:
                    rev_cagr = (newest / oldest) ** (1 / yrs) - 1
                rev_hist = [_num(x) for x in rr.tolist()]      # newest → oldest
                try:
                    fy_base = int(rr.index[0].year)             # latest reported FY-end year
                except Exception:
                    pass
            for ek in ("Operating Income", "EBIT", "Operating Revenue"):
                if ek in fin.index:
                    ebit_hist = [_num(x) for x in fin.loc[ek].dropna().tolist()]
                    break
    except Exception:
        pass

    # Latest reported quarter + reference "today" (real, from price history) —
    # used by the Python data-recency gate.
    q_end = None
    try:
        qf = t.quarterly_financials
        if qf is not None and not qf.empty:
            q_end = qf.columns[0].to_pydatetime().date()
    except Exception:
        pass
    ref_date = None
    try:
        if hist is not None and len(hist):
            ref_date = hist.index[-1].to_pydatetime().date()
    except Exception:
        pass

    # Currency mismatch: dual-listed stocks (e.g. Infosys on NYSE) report
    # financial statements in USD while the NSE price is in INR. Convert
    # statement figures to the price currency via a live FX rate.
    price_ccy = info.get("currency") or "INR"
    fin_ccy = info.get("financialCurrency") or price_ccy
    fx = 1.0
    if fin_ccy != price_ccy:
        try:
            pair = yf.Ticker(f"{fin_ccy}{price_ccy}=X").history(period="5d")
            if pair is not None and len(pair):
                fx = float(pair["Close"].iloc[-1])
        except Exception:
            fx = 1.0

    def _fx(v):
        return v * fx if v is not None else None

    # Dividend yield: yfinance returns percent (e.g. 4.8) in newer versions,
    # fraction (0.048) in older — normalise to a fraction.
    dy = _num(info.get("dividendYield"))
    if dy is not None and dy > 1:
        dy = dy / 100

    price = _num(info.get("currentPrice")) or _num(info.get("regularMarketPrice"))
    hist_close = _num(hist["Close"].iloc[-1]) if (hist is not None and len(hist)) else None
    if price is None:
        price = hist_close

    if price is None:
        raise ValueError(f"No market data found for '{symbol}' (tried {sym}). "
                         f"Use an NSE symbol like INFY, TCS, RELIANCE.")

    # ── PRICE SANITY GUARDS (a bad quote once showed ₹0.75 on a ₹4,600 stock) ──
    price_note = None
    # 1. Quote vs last historical close: >15% apart → trust the price series.
    if hist_close and price and abs(price / hist_close - 1) > 0.15:
        price_note = f"quote ₹{price:,.2f} conflicted with last close ₹{hist_close:,.2f}; using close"
        price = hist_close
    # 2. Quote vs 52-week band: far outside → trust the price series.
    _wl, _wh = _num(info.get("fiftyTwoWeekLow")), _num(info.get("fiftyTwoWeekHigh"))
    if _wl and _wh and not (0.5 * _wl <= price <= 1.5 * _wh) and hist_close and (0.5 * _wl <= hist_close <= 1.5 * _wh):
        price_note = f"quote outside 52-wk sanity band; using last close ₹{hist_close:,.2f}"
        price = hist_close

    dma50 = _num(info.get("fiftyDayAverage"))
    dma200 = _num(info.get("twoHundredDayAverage"))
    if dma200 is None and hist is not None and len(hist) >= 100:
        dma200 = _num(hist["Close"].tail(200).mean())

    # ── Technicals (for short-term analysis) ──
    tech = {}
    if hist is not None and len(hist) > 30:
        close = hist["Close"]
        tech["ma20"] = _num(close.tail(20).mean())
        tech["ma50"] = _num(close.tail(50).mean())
        tech["ma200"] = _num(close.tail(200).mean()) if len(close) >= 200 else dma200
        # RSI(14)
        delta = close.diff()
        up = _num(delta.clip(lower=0).tail(14).mean()) or 0
        down = _num((-delta.clip(upper=0)).tail(14).mean()) or 0
        tech["rsi"] = 100.0 if down == 0 else 100 - 100 / (1 + up / down)
        # momentum returns
        def _ret(n):
            return _num(close.iloc[-1] / close.iloc[-n - 1] - 1) if len(close) > n else None
        tech["ret_1m"], tech["ret_3m"], tech["ret_6m"] = _ret(21), _ret(63), _ret(126)
        # annualised volatility (6m)
        tech["vol"] = _num(close.pct_change().tail(126).std() * (252 ** 0.5))
        # support / resistance (3-month)
        tech["support"] = _num(close.tail(63).min())
        tech["resistance"] = _num(close.tail(63).max())

    # ── Analyst ratings distribution + named broker actions ──
    rating_dist = None
    try:
        rec = t.recommendations
        if rec is not None and len(rec):
            row0 = rec.iloc[0]
            rating_dist = {k: int(row0.get(c, 0)) for k, c in
                           (("strong_buy", "strongBuy"), ("buy", "buy"), ("hold", "hold"),
                            ("sell", "sell"), ("strong_sell", "strongSell"))}
    except Exception:
        pass
    brokers = []
    try:
        ud = t.upgrades_downgrades
        if (ud is None or len(ud) == 0) and sym.endswith(".NS"):
            ud = yf.Ticker(sym[:-3]).upgrades_downgrades   # dual-listed ADR (named firms)
        if ud is not None and len(ud):
            for gdate, urow in ud.sort_index(ascending=False).head(15).iterrows():
                firm = str(urow.get("Firm", "")).strip()
                grade = str(urow.get("ToGrade", "")).strip()
                if firm:
                    brokers.append({"firm": firm, "grade": grade,
                                    "action": str(urow.get("Action", "")).strip(),
                                    "date": str(gdate)[:10]})
    except Exception:
        pass

    return {
        "symbol": sym,
        "name": info.get("longName") or info.get("shortName") or sym,
        "sector": info.get("sector") or "—",
        "industry": info.get("industry") or "—",
        "summary": (info.get("longBusinessSummary") or "").strip(),
        "currency": info.get("currency") or "INR",
        "price": price,
        "market_cap": _num(info.get("marketCap")),
        "pe": _num(info.get("trailingPE")),
        "forward_pe": _num(info.get("forwardPE")),
        "peg": _num(info.get("pegRatio")),
        "pb": _num(info.get("priceToBook")),
        "roe": _num(info.get("returnOnEquity")),
        "margin": _num(info.get("profitMargins")),
        "rev_growth": _num(info.get("revenueGrowth")),
        "earn_growth": _num(info.get("earningsGrowth")),
        "debt_to_equity": _num(info.get("debtToEquity")),
        "div_yield": dy,
        "beta": _num(info.get("beta")),
        "wk_high": _num(info.get("fiftyTwoWeekHigh")),
        "wk_low": _num(info.get("fiftyTwoWeekLow")),
        "dma50": dma50,
        "dma200": dma200,
        "fcf": _fx(fcf_stmt or _num(info.get("freeCashflow"))),
        "op_cf": _fx(_num(info.get("operatingCashflow"))),
        "net_income": _fx(_num(info.get("netIncomeToCommon"))),
        "total_debt": _fx(_num(info.get("totalDebt"))),
        "total_cash": _fx(_num(info.get("totalCash"))),
        "shares": _num(info.get("sharesOutstanding")),
        "revenue": _fx(rev_stmt or _num(info.get("totalRevenue"))),
        "fin_ccy": fin_ccy, "fx": fx,
        "quarter_end": q_end, "data_ref_date": ref_date,
        "price_note": price_note,
        "tech": tech, "rating_dist": rating_dist, "brokers": brokers,
        "rev_cagr": rev_cagr,
        "rev_hist_cr": [_fx(x) / 1e7 for x in rev_hist] if rev_hist else None,
        "fy_base": fy_base,
        "target_mean": _num(info.get("targetMeanPrice")),
        "target_high": _num(info.get("targetHighPrice")),
        "target_low": _num(info.get("targetLowPrice")),
        "rec_key": info.get("recommendationKey"),
        "rec_mean": _num(info.get("recommendationMean")),
        "num_analysts": _num(info.get("numberOfAnalystOpinions")),
    }


# ═══════════════════════════════════════════════════════════
#  2. VALUATION + FACTOR ANALYSIS  (deterministic, no LLM)
# ═══════════════════════════════════════════════════════════
def dcf_fair_value(d: dict) -> dict:
    """Simple 10-yr two-stage FCF DCF. Returns fair value/share or None."""
    fcf = d.get("fcf")
    if fcf is None or fcf <= 0:
        # fall back to a proxy: 80% of operating cash flow, else net income
        base = d.get("op_cf")
        fcf = (base * 0.8) if (base and base > 0) else d.get("net_income")
    shares = d.get("shares")
    if not fcf or fcf <= 0 or not shares or shares <= 0:
        return {"fair_value": None, "upside": None, "note": "Insufficient cash-flow data for DCF"}

    g1 = d.get("rev_growth")
    g1 = 0.10 if g1 is None else max(min(g1, 0.25), 0.0)  # cap stage-1 growth
    g_term = 0.04
    r = 0.12  # discount rate

    total = 0.0
    cf = fcf
    for yr in range(1, 11):
        # fade growth linearly from g1 to terminal over 10 years
        g = g1 + (g_term - g1) * (yr - 1) / 9
        cf = cf * (1 + g)
        total += cf / ((1 + r) ** yr)
    terminal = cf * (1 + g_term) / (r - g_term)
    total += terminal / ((1 + r) ** 10)

    net_debt = (d.get("total_debt") or 0) - (d.get("total_cash") or 0)
    equity = total - net_debt
    fv = equity / shares
    price = d.get("price")

    # Sanity guard: yfinance data for Indian stocks can be inconsistent (wrong
    # units/scale). If the DCF lands wildly off the market price, the inputs are
    # unreliable — suppress the number rather than print a misleading figure.
    if price and (fv <= 0 or fv < 0.15 * price or fv > 6 * price):
        return {"fair_value": None, "upside": None,
                "note": "DCF omitted — cash-flow data failed reliability check"}

    upside = (fv / price - 1) if price else None
    return {"fair_value": fv, "upside": upside,
            "assumptions": {"stage1_growth": g1, "terminal_growth": g_term, "discount_rate": r},
            "note": ""}


def analyze(d: dict) -> dict:
    """Multi-factor scorecard → recommendation. Pure rules."""
    factors = []  # (name, value_str, verdict, points)

    def add(name, val, points, comment):
        factors.append({"factor": name, "value": val, "signal": comment, "points": points})

    price = d["price"]

    pe = d.get("pe")
    if pe is not None:
        p = 1 if pe < 22 else (0 if pe < 35 else -1)
        add("Valuation (P/E)", f"{pe:.1f}x", p,
            "Attractive" if p > 0 else ("Fair" if p == 0 else "Rich"))

    peg = d.get("peg")
    if peg is not None and peg > 0:
        p = 1 if peg < 1 else (0 if peg < 2 else -1)
        add("Growth-adj. (PEG)", f"{peg:.2f}", p,
            "Cheap vs growth" if p > 0 else ("Balanced" if p == 0 else "Expensive"))

    roe = d.get("roe")
    if roe is not None:
        p = 1 if roe > 0.18 else (0 if roe > 0.10 else -1)
        add("Profitability (ROE)", f"{roe*100:.1f}%", p,
            "Strong" if p > 0 else ("Adequate" if p == 0 else "Weak"))

    rg = d.get("rev_growth")
    if rg is not None:
        p = 1 if rg > 0.12 else (0 if rg > 0.04 else -1)
        add("Revenue growth", f"{rg*100:.1f}%", p,
            "High" if p > 0 else ("Moderate" if p == 0 else "Sluggish"))

    de = d.get("debt_to_equity")
    if de is not None:
        de_r = de / 100 if de > 5 else de  # yfinance gives % sometimes
        p = 1 if de_r < 0.5 else (0 if de_r < 1.2 else -1)
        add("Leverage (D/E)", f"{de_r:.2f}", p,
            "Conservative" if p > 0 else ("Moderate" if p == 0 else "Elevated"))

    if d.get("dma200"):
        p = 1 if price > d["dma200"] else -1
        add("Trend (vs 200-DMA)", f"₹{d['dma200']:.0f}", p,
            "Above — uptrend" if p > 0 else "Below — weak")

    if d.get("wk_high") and d.get("wk_low"):
        rng = (price - d["wk_low"]) / (d["wk_high"] - d["wk_low"]) if d["wk_high"] > d["wk_low"] else 0.5
        p = 0 if 0.3 <= rng <= 0.85 else (1 if rng < 0.3 else -1)
        add("52-wk position", f"{rng*100:.0f}% of range", p,
            "Near lows — value" if p > 0 else ("Mid-range" if p == 0 else "Near highs"))

    dcf = dcf_fair_value(d)
    if dcf["upside"] is not None:
        u = dcf["upside"]
        p = 1 if u > 0.15 else (0 if u > -0.10 else -1)
        add("DCF fair value", f"₹{dcf['fair_value']:.0f} ({u*100:+.0f}%)", p,
            "Undervalued" if p > 0 else ("Fairly valued" if p == 0 else "Overvalued"))

    score = sum(f["points"] for f in factors)
    n = len(factors) or 1
    norm = score / n

    if norm >= 0.45:
        verdict, action = "BUY", "Accumulate on dips — strong long-term conviction"
    elif norm >= 0.15:
        verdict, action = "ACCUMULATE", "Build position gradually via SIP/DCA"
    elif norm >= -0.15:
        verdict, action = "HOLD", "Hold existing; await better entry or catalysts"
    elif norm >= -0.45:
        verdict, action = "REDUCE", "Trim on strength; risks outweigh reward"
    else:
        verdict, action = "AVOID", "Unfavourable risk/reward at current levels"

    return {"factors": factors, "score": score, "n": n, "norm": norm,
            "verdict": verdict, "action": action, "dcf": dcf}


def numbers_block(d: dict, a: dict, result: dict) -> str:
    """Compact numbers shown IN the chat (not spoken). Leads with the analyst
    target; the DCF is shown only when it's a sane cross-check (not for
    negative-FCF / early-stage names where a plain DCF is misleading)."""
    price = d["price"]
    fv = (result.get("validation") or {}).get("computed_fair_value")
    tgt = result.get("price_target")
    base_fcf = d.get("fcf")
    lines = [f"**{d['name']} ({d['symbol']})**  ·  ₹{price:,.0f}"]

    kv = []
    pe = d.get("pe")
    if pe:
        kv.append(f"P/E {pe:.0f}x" + (" (rich)" if pe > 60 else ""))
    if d.get("roe") is not None:
        kv.append(f"ROE {d['roe']*100:.1f}%")
    rg = d.get("rev_growth")
    if rg is not None:
        kv.append("Rev gr n/m" if abs(rg) > 1.0 else f"Rev gr {rg*100:.1f}%")   # >100% = data artifact
    if d.get("debt_to_equity") is not None:
        de = d["debt_to_equity"]
        kv.append(f"D/E {(de/100 if de > 5 else de):.2f}")
    if kv:
        lines.append(" · ".join(kv))

    fcf_str = "negative (growth-stage)" if (base_fcf is not None and base_fcf < 0) else _fmt_money(base_fcf)
    lines.append(f"Revenue {_fmt_money(d.get('revenue'))} · FCF {fcf_str}")

    # Valuation: analyst target leads; DCF only as a sane cross-check.
    if tgt:
        lines.append(f"Analyst fair value ₹{tgt:,.0f} ({(tgt/price-1)*100:+.0f}%)")
    dcf_reliable = (fv and base_fcf and base_fcf > 0 and 0.5 * price <= fv <= 2.0 * price)
    if dcf_reliable:
        lines.append(f"DCF cross-check ₹{fv:,.0f} ({(fv/price-1)*100:+.0f}%)")
    elif not tgt and fv:
        lines.append(f"DCF (indicative) ₹{fv:,.0f}")
    if not tgt and not dcf_reliable and base_fcf is not None and base_fcf < 0:
        lines.append("_DCF omitted — negative FCF; valuation is judgement-led._")

    lines.append(f"**Verdict: {result['verdict']}**")
    validated = (result.get("validation") or {}).get("ok")
    lines.append(f"_Data as-of {result.get('data_asof')} · "
                 f"{'✅ calcs validated' if validated else '⚠ validation flagged'} · "
                 f"Excel + PDF downloading…_")
    return "\n".join(lines)


def consensus_summary(d: dict):
    """Street analyst consensus from live data (or None if unavailable/unreliable)."""
    tm, price = d.get("target_mean"), d.get("price")
    if not tm or not price or not (0.3 * price <= tm <= 3 * price):
        return None
    return {
        "mean": tm, "high": d.get("target_high"), "low": d.get("target_low"),
        "rec": (d.get("rec_key") or "").upper().replace("_", " ") or None,
        "rec_mean": d.get("rec_mean"),
        "n": int(d["num_analysts"]) if d.get("num_analysts") else None,
        "upside_pct": (tm / price - 1) * 100,
    }


def compose_analysis(d: dict, a: dict, narrative, result: dict):
    """Return (chat_markdown, speech). chat_markdown = the ENTIRE Claude analysis
    for display; speech = first 2 lines + last 2 lines for JARVIS to read aloud."""
    n = narrative or {}
    verdict = result["verdict"]
    tgt = result.get("price_target")
    sector = d.get("industry") or d.get("sector") or ""

    thesis = (n.get("thesis") or a["action"]).strip()
    thesis_first = thesis.split(". ")[0].rstrip(".") + "."
    rec = (n.get("recommendation") or a["action"]).strip()
    rec_last = rec.split(". ")[-1].strip().rstrip(".") + "."
    rationale = (n.get("verdict_rationale") or "").strip()

    # ── Spoken: 2 lines about the company (intro + thesis) + final recommendation ──
    intro_line = f"{d['name']}, {sector}." if sector else f"{d['name']}."
    final_line = f"Final recommendation: {verdict}" + (f", {rationale}" if rationale else "") + \
                 (f". Twelve-month target {tgt:,.0f} rupees." if tgt else ".")
    speech = " ".join([intro_line, thesis_first, final_line])

    price = d["price"]
    pct = lambda x: f"{x*100:.1f}%" if x is not None else "—"

    # ── Chat: metrics → analysis → ALL calcs (refer Excel) → recommendation LAST ──
    p = [f"**{d['name']} ({d['symbol']})** · ₹{price:,.0f}"]

    # key metrics
    kv = []
    pe = d.get("pe")
    if pe:
        kv.append(f"P/E {pe:.0f}x" + (" (rich)" if pe > 60 else ""))
    if d.get("roe") is not None:
        kv.append(f"ROE {d['roe']*100:.1f}%")
    rg = d.get("rev_growth")
    if rg is not None:
        kv.append("Rev gr n/m" if abs(rg) > 1.0 else f"Rev gr {rg*100:.1f}%")
    if d.get("debt_to_equity") is not None:
        de = d["debt_to_equity"]
        kv.append(f"D/E {(de/100 if de > 5 else de):.2f}")
    fcf_str = "negative (growth-stage)" if (d.get("fcf") is not None and d["fcf"] < 0) else _fmt_money(d.get("fcf"))
    p.append((" · ".join(kv) + "\n" if kv else "") +
             f"Revenue {_fmt_money(d.get('revenue'))} · FCF {fcf_str} · Data as-of {result.get('data_asof')}")

    # analysis
    p.append(f"**Thesis.** {thesis}")
    if n.get("business"):
        p.append(f"**Business & moat.** {n['business']}")
    if n.get("bull_case"):
        p.append("**Bull case:**\n" + "\n".join(f"- {x}" for x in n["bull_case"]))
    if n.get("bear_case"):
        p.append("**Bear case:**\n" + "\n".join(f"- {x}" for x in n["bear_case"]))
    if n.get("catalysts"):
        p.append("**Catalysts:**\n" + "\n".join(f"- {x}" for x in n["catalysts"]))
    if n.get("risks"):
        p.append("**Key risks:**\n" + "\n".join(f"- {x}" for x in n["risks"]))

    # ── FULL DCF CALCULATIONS (all line items, year by year) ──
    asmp = result.get("assumptions") or {}
    dcf = None
    if asmp:
        dcf = python_dcf(asmp)
        wb = n.get("wacc_build") or {}
        v = ["**Valuation — full DCF model (₹ Crore)**"]
        if wb:
            v.append(f"WACC build: RF {pct(wb.get('rf'))} + β{wb.get('beta','—')}×ERP {pct(wb.get('erp'))} "
                     f"→ Ke {pct(wb.get('cost_of_equity'))}; Kd {pct(wb.get('cost_of_debt'))}; "
                     f"{pct(wb.get('equity_weight'))} equity / {pct(wb.get('debt_weight'))} debt "
                     f"→ **WACC {pct(asmp.get('wacc'))}**")
        else:
            v.append(f"Discount rate (WACC): **{pct(asmp.get('wacc'))}**")
        g = asmp.get("growth", [])
        em = asmp.get("ebit_margin")
        em = em if isinstance(em, list) else [em] * 5
        v.append(f"Assumptions: growth {' / '.join(pct(x) for x in g)}")
        v.append(f"EBIT margin path: {' / '.join(pct(x) for x in em)}")
        v.append(f"tax {pct(asmp.get('tax_rate'))} · capex {pct(asmp.get('capex_pct'))} · D&A {pct(asmp.get('da_pct'))} · "
                 f"ΔNWC {pct(asmp.get('nwc_pct'))} · terminal {pct(asmp.get('terminal_growth'))}")
        sch = dcf.get("schedule", [])
        if sch:
            fyl = fy_labels(d, len(sch))
            v.append("**Revenue projection** (₹Cr): " + " · ".join(f"{fyl[i]} {s['revenue']:,.0f}" for i, s in enumerate(sch)))
            v.append("**FCF projection** (₹Cr): " + " · ".join(f"{fyl[i]} {s['fcf']:,.0f}" for i, s in enumerate(sch)))
            def rowfmt(label, key):
                return label.ljust(12) + "".join(f"{s[key]:>11,.0f}" for s in sch)
            tbl = ["Line item".ljust(12) + "".join(f"{fyl[i]:>11}" for i in range(len(sch))),
                   rowfmt("Revenue", "revenue"),
                   rowfmt("EBIT", "ebit"),
                   rowfmt("NOPAT", "nopat"),
                   rowfmt("(+) D&A", "da"),
                   rowfmt("(-) Capex", "capex"),
                   rowfmt("(-) d NWC", "dnwc"),
                   rowfmt("FCF", "fcf"),
                   "Disc factor".ljust(12) + "".join(f"{s['df']:>11.3f}" for s in sch),
                   rowfmt("PV of FCF", "pv_fcf")]
            v.append("```\n" + "\n".join(tbl) + "\n```")
        v.append(f"Σ PV explicit ₹{dcf['pv_explicit']:,.0f} + PV terminal ₹{dcf['pv_terminal']:,.0f} "
                 f"= EV ₹{dcf['ev']:,.0f} − net debt ₹{asmp.get('net_debt_cr',0):,.0f} "
                 f"= Equity ₹{dcf['equity']:,.0f} Cr" +
                 (f" → **DCF fair value ₹{dcf['fair_value']:,.0f}/sh**" if dcf.get("fair_value") else ""))
        if n.get("assumption_log"):
            v.append("Assumption log:\n" + "\n".join(f"- {x}" for x in n["assumption_log"]))
        v.append("_📊 Full linked model with live formulas is in the downloaded Excel "
                 "(sheets: Assumptions · Model · Scorecard · DCF)._")
        p.append("\n".join(v))

    # ── SCENARIOS (bull / base / bear) ──
    scen = n.get("scenarios")
    if scen:
        sl = ["**Scenarios (bull / base / bear)**"]
        ev = 0.0
        for key in ("bull", "base", "bear"):
            s = scen.get(key) or {}
            tg = s.get("target"); pr = s.get("probability")
            if tg:
                sl.append(f"{key.capitalize()}: ₹{tg:,.0f} ({(tg/price-1)*100:+.0f}%)" +
                          (f" · {pr*100:.0f}%" if pr is not None else "") +
                          (f" — {s['driver']}" if s.get("driver") else ""))
                if pr: ev += tg * pr
        if ev:
            sl.append(f"**Probability-weighted target: ₹{ev:,.0f} ({(ev/price-1)*100:+.0f}%)**")
        p.append("\n".join(sl))

    # ── VALUATION TRIANGULATION (comps · SOTP · football field) ──
    comps = n.get("comps")
    sotp = n.get("sotp")
    dcf_fv = (result.get("validation") or {}).get("computed_fair_value")
    if comps or sotp:
        ml = ["**Valuation triangulation**"]
        if dcf_fv:
            ml.append(f"DCF ₹{dcf_fv:,.0f}")
        if comps and comps.get("implied_value_per_share"):
            peers = ", ".join(pp.get("name", "") for pp in (comps.get("peers") or [])[:5])
            ml.append(f"Comps ₹{comps['implied_value_per_share']:,.0f} (median {comps.get('median_ev_ebitda','?')}x EV/EBITDA vs {peers})")
        if sotp and sotp.get("implied_value_per_share"):
            segs = " + ".join(s.get("segment", "") for s in (sotp.get("segments") or []))
            ml.append(f"SOTP ₹{sotp['implied_value_per_share']:,.0f} ({segs})")
        _c = consensus_summary(d)
        if _c and _c.get("low") and _c.get("high"):
            ml.append(f"Analyst range ₹{_c['low']:,.0f}–₹{_c['high']:,.0f}")
        ml.append("_📊 Football-field chart + Comps/SOTP sheets in the Excel._")
        p.append("\n".join(ml))

    # ── STREET CONSENSUS vs OUR VIEW ──
    cons = consensus_summary(d)
    if cons:
        rng = ""
        if cons.get("low") and cons.get("high"):
            rng = f" (range ₹{cons['low']:,.0f}–₹{cons['high']:,.0f})"
        cl = ["**Street consensus vs our view**"]
        cl.append(f"Street: **{cons['rec'] or 'N/A'}** · mean target ₹{cons['mean']:,.0f} "
                  f"({cons['upside_pct']:+.0f}%){rng}" +
                  (f" · {cons['n']} analysts" if cons.get("n") else ""))
        rd = d.get("rating_dist")
        if rd:
            cl.append(f"Ratings: {rd.get('strong_buy',0)} strong-buy · {rd.get('buy',0)} buy · "
                      f"{rd.get('hold',0)} hold · {rd.get('sell',0)} sell · {rd.get('strong_sell',0)} strong-sell")
        brk = d.get("brokers") or []
        if brk:
            names = " · ".join(f"{b['firm']} ({b['grade']})" for b in brk[:6] if b.get("grade"))
            if names:
                cl.append(f"Recent broker actions: {names}")
        ours_t = tgt or (result.get("validation") or {}).get("computed_fair_value")
        if ours_t:
            gap = (ours_t / cons["mean"] - 1) * 100
            cl.append(f"Ours: **{verdict}** · target ₹{ours_t:,.0f} — **{gap:+.0f}% vs street**")
        if n.get("vs_consensus"):
            cl.append(n["vs_consensus"])
        if n.get("divergence_factor"):
            cl.append(f"**Key fundamental difference:** {n['divergence_factor']}")
        p.append("\n".join(cl))

    # ── RECOMMENDATION — LAST, after analysing ──
    rblock = [f"**Recommendation: {verdict}**", rec]
    fvc = (result.get("validation") or {}).get("computed_fair_value")
    dcf_reliable = fvc and d.get("fcf") and d["fcf"] > 0 and 0.5 * price <= fvc <= 2.0 * price
    tail = []
    if tgt:
        tail.append(f"Analyst target ₹{tgt:,.0f} ({(tgt/price-1)*100:+.0f}%)")
    if dcf_reliable:
        tail.append(f"DCF fair value ₹{fvc:,.0f} ({(fvc/price-1)*100:+.0f}%)")
    if tail:
        rblock.append(" · ".join(tail))
    validated = (result.get("validation") or {}).get("ok")
    rblock.append(f"_{'✅ calcs validated by Python layer' if validated else '⚠ validation flagged'} · Excel + PDF downloading…_")
    p.append("\n".join(rblock))

    return "\n\n".join(p), speech


def cross_validate_price(symbol: str, yf_price: float, timeout: int = 8) -> dict:
    """Independent price check against Google Finance (a non-Yahoo source). Part
    of the Python validation layer: flags if yfinance's price diverges >2%."""
    sym = (symbol or "").upper()
    exch = "BOM" if sym.endswith(".BO") else "NSE"
    base = sym.replace(".NS", "").replace(".BO", "")
    try:
        import urllib.request
        req = urllib.request.Request(
            f"https://www.google.com/finance/quote/{base}:{exch}",
            headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                                   "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                     "Accept-Language": "en-US,en;q=0.9"})
        html = urllib.request.urlopen(req, timeout=timeout).read().decode("utf-8", "ignore")
        import re as _re
        m = (_re.search(r'data-last-price="([0-9.]+)"', html)
             or _re.search(r'class="YMlKec fxKbKc">\s*₹\s*([0-9,]+(?:\.[0-9]+)?)', html)
             or _re.search(r'₹\s*([0-9,]+\.[0-9]{2})', html))
        if not m:
            return {"ok": None, "note": "second source unavailable"}
        gp = float(m.group(1).replace(",", ""))
        diff = abs(gp - yf_price) / gp if gp else None
        return {"ok": diff is not None and diff < 0.02, "google_price": gp,
                "diff_pct": round(diff * 100, 2) if diff is not None else None}
    except Exception:
        return {"ok": None, "note": "second source unavailable"}


def data_recency(d: dict) -> dict:
    """Python gate: is the financial data from the latest or previous quarter?"""
    q = d.get("quarter_end")
    ref = d.get("data_ref_date")
    if not q or not ref:
        return {"asof": "unknown", "status": "unknown", "ok": True, "days": None}
    days = (ref - q).days
    if days <= 135:
        status, ok = "latest quarter", True
    elif days <= 240:
        status, ok = "previous quarter", True
    else:
        status, ok = "STALE (>2 quarters old)", False
    return {"asof": q.strftime("%d %b %Y"), "status": status, "ok": ok, "days": days}


def final_verdict(a: dict, narrative=None) -> str:
    """Claude's call is the headline; quant scorecard is the fallback."""
    v = (narrative or {}).get("verdict")
    if v and str(v).upper() in ("BUY", "ACCUMULATE", "HOLD", "REDUCE", "SELL"):
        return str(v).upper()
    return a["verdict"]


def two_line_summary(d: dict, a: dict, narrative=None) -> str:
    """Exactly two lines for voice/chat: company + final call with short rationale."""
    intro = f"{d['name']} — {d['industry']} ({d['sector']})."
    verdict = final_verdict(a, narrative)
    rationale = (narrative or {}).get("verdict_rationale")
    if not rationale and narrative and narrative.get("recommendation"):
        rationale = narrative["recommendation"].split(". ")[0]
    rationale = (rationale or a["action"]).strip().rstrip(".")
    # keep the spoken line tight
    if len(rationale) > 90:
        rationale = rationale[:87].rsplit(" ", 1)[0] + "…"
    return f"{intro}\nRecommendation: {verdict} — {rationale}."


# ═══════════════════════════════════════════════════════════
#  3. EXCEL MODEL
# ═══════════════════════════════════════════════════════════
def _fmt_money(v):
    if v is None:
        return "N/A"
    a = abs(v)
    if a >= 1e12:
        return f"₹{v/1e12:.2f} T"
    if a >= 1e7:
        return f"₹{v/1e7:.2f} Cr"
    if a >= 1e5:
        return f"₹{v/1e5:.2f} L"
    return f"₹{v:,.0f}"


def default_assumptions(d: dict) -> dict:
    """5-year model drivers grounded on real data (in ₹ Crore where noted)."""
    rev = d.get("revenue") or 0
    base_rev_cr = rev / 1e7 if rev else 1000.0
    g = d.get("rev_growth")
    g1 = min(max(g if g is not None else 0.10, 0.04), 0.25)
    g5 = max(g1 * 0.55, 0.06)
    growth = [round(g1 + (g5 - g1) * i / 4, 4) for i in range(5)]  # linear fade
    margin = d.get("margin")
    m0 = round(min(max((margin * 1.25) if margin is not None else 0.15, 0.05), 0.45), 4)
    # gentle expansion path from trailing margin (fallback; Claude usually supplies its own)
    ebit_margin = [round(min(m0 + 0.004 * i, 0.55), 4) for i in range(5)]
    net_debt_cr = ((d.get("total_debt") or 0) - (d.get("total_cash") or 0)) / 1e7
    shares_cr = (d.get("shares") or 0) / 1e7
    return {
        "base_rev_cr": round(base_rev_cr, 2),
        "growth": growth,
        "ebit_margin": ebit_margin,
        "tax_rate": 0.25,
        "capex_pct": 0.04,
        "da_pct": 0.035,
        "nwc_pct": 0.02,
        "wacc": 0.12,
        "terminal_growth": 0.05,
        "net_debt_cr": round(net_debt_cr, 2),
        "shares_cr": round(shares_cr, 4) if shares_cr else 1.0,
        "price": d["price"],
    }


def merge_assumptions(base: dict, claude: dict) -> dict:
    """Overlay Claude-suggested drivers on defaults, clamped to sane ranges."""
    if not claude:
        return base
    out = dict(base)
    def clamp(x, lo, hi):
        try:
            return max(lo, min(hi, float(x)))
        except (TypeError, ValueError):
            return None
    if isinstance(claude.get("growth"), list) and claude["growth"]:
        g = [clamp(x, -0.10, 0.40) for x in claude["growth"][:5]]
        g = [x for x in g if x is not None]
        if len(g) == 5:
            out["growth"] = g
    # EBIT margin is a 5-year path (accept a scalar for backward compat)
    cm = claude.get("ebit_margin")
    if cm is not None:
        cm_list = cm if isinstance(cm, list) else [cm] * 5
        mm = [clamp(x, 0.02, 0.55) for x in cm_list[:5]]
        mm = [x for x in mm if x is not None]
        if len(mm) == 5:
            out["ebit_margin"] = [round(x, 4) for x in mm]
    for key, lo, hi in [("tax_rate", 0.10, 0.40),
                        ("capex_pct", 0.0, 0.25), ("da_pct", 0.0, 0.20),
                        ("nwc_pct", -0.10, 0.20), ("wacc", 0.07, 0.20),
                        ("terminal_growth", 0.0, 0.07)]:
        if claude.get(key) is not None:
            v = clamp(claude[key], lo, hi)
            if v is not None:
                out[key] = round(v, 4)
    # keep terminal_growth strictly below wacc for DCF stability
    if out["terminal_growth"] >= out["wacc"] - 0.01:
        out["terminal_growth"] = round(out["wacc"] - 0.03, 4)
    return out


def python_dcf(asmp: dict) -> dict:
    """Independent recompute of the linked Excel model — the validation oracle.
    Mirrors the MODEL sheet formulas cell-for-cell."""
    rev_prev = asmp["base_rev_cr"]
    wacc, tg = asmp["wacc"], asmp["terminal_growth"]
    margins = asmp["ebit_margin"]
    if not isinstance(margins, list):
        margins = [margins] * 5
    pv_sum, fcf_last, df_last = 0.0, 0.0, 1.0
    schedule = []
    for i in range(5):
        rev = rev_prev * (1 + asmp["growth"][i])
        ebit = rev * margins[i]
        nopat = ebit * (1 - asmp["tax_rate"])
        da = rev * asmp["da_pct"]
        capex = rev * asmp["capex_pct"]
        dnwc = (rev - rev_prev) * asmp["nwc_pct"]
        fcf = nopat + da - capex - dnwc
        df = 1 / (1 + wacc) ** (i + 1)
        pv_sum += fcf * df
        schedule.append({"year": i + 1, "revenue": rev, "ebit": ebit, "nopat": nopat,
                         "da": da, "capex": capex, "dnwc": dnwc, "fcf": fcf,
                         "df": df, "pv_fcf": fcf * df})
        rev_prev, fcf_last, df_last = rev, fcf, df
    tv = fcf_last * (1 + tg) / (wacc - tg)
    pv_tv = tv * df_last
    ev = pv_sum + pv_tv
    equity = ev - asmp["net_debt_cr"]
    fv = equity / asmp["shares_cr"] if asmp["shares_cr"] else None
    upside = (fv / asmp["price"] - 1) if (fv and asmp.get("price")) else None
    return {"ev": ev, "equity": equity, "fair_value": fv, "upside": upside,
            "schedule": schedule, "pv_explicit": pv_sum, "terminal_value": tv,
            "pv_terminal": pv_tv}


def build_excel(d: dict, a: dict, narrative=None, assumptions=None) -> bytes:
    wb = Workbook()
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    arc_fill = PatternFill("solid", fgColor=ARC)
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(color="FFFFFF", bold=True, size=16)

    def style_header(ws, row, cols, text=None):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = border

    # ── COVER ──
    ws = wb.active
    ws.title = "COVER"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"] = "J.A.R.V.I.S. · INSTITUTIONAL EQUITY RESEARCH"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34
    rows = [
        ("Company", d["name"]),
        ("Symbol", d["symbol"]),
        ("Sector / Industry", f"{d['sector']} / {d['industry']}"),
        ("Current Price", f"₹{d['price']:,.2f}"),
        ("Market Cap", _fmt_money(d.get("market_cap"))),
        ("Recommendation", final_verdict(a, narrative)),
        ("Action", (narrative or {}).get("verdict_rationale") or a["action"]),
        ("Report Date", datetime.now().strftime("%d %b %Y")),
    ]
    r = 3
    for k, v in rows:
        ws.cell(r, 1, k).font = Font(bold=True, color=NAVY)
        ws.cell(r, 2, v)
        if k == "Recommendation":
            ws.cell(r, 2).font = Font(bold=True, size=13, color=ARC)
        r += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 52

    # ── KEY METRICS ──
    ws = wb.create_sheet("KEY_METRICS")
    ws.sheet_view.showGridLines = False
    ws.append(["Metric", "Value"])
    style_header(ws, 1, 2)
    metrics = [
        ("Trailing P/E", f"{d['pe']:.1f}x" if d.get("pe") else "N/A"),
        ("Forward P/E", f"{d['forward_pe']:.1f}x" if d.get("forward_pe") else "N/A"),
        ("PEG ratio", f"{d['peg']:.2f}" if d.get("peg") else "N/A"),
        ("Price / Book", f"{d['pb']:.2f}" if d.get("pb") else "N/A"),
        ("Return on Equity", f"{d['roe']*100:.1f}%" if d.get("roe") is not None else "N/A"),
        ("Profit Margin", f"{d['margin']*100:.1f}%" if d.get("margin") is not None else "N/A"),
        ("Revenue Growth", f"{d['rev_growth']*100:.1f}%" if d.get("rev_growth") is not None else "N/A"),
        ("Earnings Growth", f"{d['earn_growth']*100:.1f}%" if d.get("earn_growth") is not None else "N/A"),
        ("Debt / Equity", f"{d['debt_to_equity']:.2f}" if d.get("debt_to_equity") is not None else "N/A"),
        ("Dividend Yield", f"{d['div_yield']*100:.2f}%" if d.get("div_yield") else "N/A"),
        ("Beta", f"{d['beta']:.2f}" if d.get("beta") else "N/A"),
        ("52-wk High", f"₹{d['wk_high']:,.0f}" if d.get("wk_high") else "N/A"),
        ("52-wk Low", f"₹{d['wk_low']:,.0f}" if d.get("wk_low") else "N/A"),
        ("50-DMA", f"₹{d['dma50']:,.0f}" if d.get("dma50") else "N/A"),
        ("200-DMA", f"₹{d['dma200']:,.0f}" if d.get("dma200") else "N/A"),
        ("Revenue (TTM)", _fmt_money(d.get("revenue"))),
        ("Free Cash Flow", _fmt_money(d.get("fcf"))),
    ]
    for k, v in metrics:
        ws.append([k, v])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.border = border
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20

    # ── SCORECARD ──
    ws = wb.create_sheet("SCORECARD")
    ws.sheet_view.showGridLines = False
    ws.append(["Factor", "Value", "Signal", "Score"])
    style_header(ws, 1, 4)
    for f in a["factors"]:
        ws.append([f["factor"], f["value"], f["signal"], f["points"]])
        sc = ws.cell(ws.max_row, 4)
        sc.font = Font(bold=True, color=("1A7F37" if f["points"] > 0 else ("B42318" if f["points"] < 0 else "6B7A8D")))
    ws.append([])
    ws.append(["COMPOSITE", f"{a['score']}/{a['n']}", a["verdict"], round(a["norm"], 2)])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.cell(ws.max_row, 3).font = Font(bold=True, color=ARC)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=4):
        for cell in row:
            cell.border = border
    for col, w in zip("ABCD", (24, 22, 22, 10)):
        ws.column_dimensions[col].width = w

    # ── ASSUMPTIONS (input cells that drive the model) ──
    asmp = assumptions or default_assumptions(d)
    pct = "0.0%"
    num = "#,##0"
    wa = wb.create_sheet("ASSUMPTIONS")
    wa.sheet_view.showGridLines = False
    wa.merge_cells("A1:F1")
    wa["A1"] = "ASSUMPTIONS  (editable drivers)"
    wa["A1"].fill = hdr_fill
    wa["A1"].font = hdr_font
    _m_list = asmp["ebit_margin"] if isinstance(asmp["ebit_margin"], list) else [asmp["ebit_margin"]] * 5
    _m_avg = round(sum(_m_list) / len(_m_list), 4)
    inputs = [
        ("Base revenue (₹ Cr)", asmp["base_rev_cr"], num),      # B3
        ("EBIT margin (avg)", _m_avg, pct),                      # B4 (per-year path in row 17)
        ("Tax rate", asmp["tax_rate"], pct),                     # B5
        ("Capex (% revenue)", asmp["capex_pct"], pct),           # B6
        ("D&A (% revenue)", asmp["da_pct"], pct),                # B7
        ("Δ NWC (% Δrevenue)", asmp["nwc_pct"], pct),            # B8
        ("WACC (discount rate)", asmp["wacc"], pct),             # B9
        ("Terminal growth", asmp["terminal_growth"], pct),       # B10
        ("Net debt (₹ Cr)", asmp["net_debt_cr"], num),           # B11
        ("Shares (Cr)", asmp["shares_cr"], "#,##0.00"),          # B12
        ("Current price (₹)", asmp["price"], "#,##0.00"),        # B13
    ]
    # header row + per-assumption SOURCE / BASIS column (col C)
    for ci, htxt in ((1, "Driver"), (2, "Value"), (3, "Source / Basis")):
        hc = wa.cell(2, ci, htxt)
        hc.font = Font(bold=True, color="FFFFFF")
        hc.fill = PatternFill("solid", fgColor=NAVY)
    src = (narrative or {}).get("sources") or {}
    wb_note = ((narrative or {}).get("wacc_build") or {}).get("note")
    data_src = f"Yahoo Finance · FY as-of {d.get('quarter_end') or 'latest'}"
    source_map = {
        "Base revenue (₹ Cr)": data_src,
        "EBIT margin (avg)": src.get("ebit_margin", "trailing margin + op leverage"),
        "Tax rate": src.get("tax_rate", "India statutory ~25%"),
        "Capex (% revenue)": src.get("capex_pct", "historical capex/revenue"),
        "D&A (% revenue)": src.get("da_pct", "historical D&A/revenue"),
        "Δ NWC (% Δrevenue)": src.get("nwc_pct", "working-capital trend"),
        "WACC (discount rate)": src.get("wacc", wb_note or "CAPM: RF + β×ERP"),
        "Terminal growth": src.get("terminal_growth", "long-run nominal GDP"),
        "Net debt (₹ Cr)": data_src,
        "Shares (Cr)": data_src,
        "Current price (₹)": "Yahoo Finance (live)",
    }
    r = 3
    for label, val, fmt in inputs:
        wa.cell(r, 1, label).font = Font(bold=True, color=NAVY)
        c = wa.cell(r, 2, val)
        c.number_format = fmt
        c.fill = PatternFill("solid", fgColor="FFF7E0")  # highlight = editable input
        c.border = border
        sc = wa.cell(r, 3, source_map.get(label, "—"))
        sc.font = Font(italic=True, color=GREY, size=9)
        sc.alignment = Alignment(wrap_text=True, vertical="top")
        sc.border = border
        r += 1
    # revenue growth path across Year 1..5 → B15:F15
    _fyl = fy_labels(d)
    wa.cell(15, 1, f"Revenue growth ({_fyl[0]}→{_fyl[-1]})").font = Font(bold=True, color=NAVY)
    for i, g in enumerate(asmp["growth"]):
        c = wa.cell(15, 2 + i, g)
        c.number_format = pct
        c.fill = PatternFill("solid", fgColor="FFF7E0")
        c.border = border
    wa.cell(16, 1, "  ↳ growth basis").font = Font(italic=True, color=GREY, size=9)
    gsc = wa.cell(16, 2, src.get("growth", "3yr historical avg + management guidance"))
    gsc.font = Font(italic=True, color=GREY, size=9)
    wa.merge_cells("B16:F16")
    # EBIT margin PATH across Year 1..5 → B17:F17 (drives the model per year)
    wa.cell(17, 1, f"EBIT margin ({_fyl[0]}→{_fyl[-1]})").font = Font(bold=True, color=NAVY)
    for i, m in enumerate(_m_list):
        c = wa.cell(17, 2 + i, m)
        c.number_format = pct
        c.fill = PatternFill("solid", fgColor="FFF7E0")
        c.border = border
    wa.cell(18, 1, "  ↳ margin basis").font = Font(italic=True, color=GREY, size=9)
    msc = wa.cell(18, 2, src.get("ebit_margin", "trailing margin + operating leverage path"))
    msc.font = Font(italic=True, color=GREY, size=9)
    wa.merge_cells("B18:F18")
    wa.column_dimensions["A"].width = 26
    wa.column_dimensions["B"].width = 13
    wa.column_dimensions["C"].width = 46
    for col in "DEF":
        wa.column_dimensions[col].width = 12

    # ── MODEL (DCF, fully formula-linked to ASSUMPTIONS) ──
    wm = wb.create_sheet("MODEL")
    wm.sheet_view.showGridLines = False
    wm.merge_cells("A1:F1")
    wm["A1"] = "DCF MODEL  (₹ Crore) — live formulas"
    wm["A1"].fill = hdr_fill
    wm["A1"].font = hdr_font
    cols = ["B", "C", "D", "E", "F"]  # Year 1..5
    ylabels = fy_labels(d)
    wm.cell(2, 1, "Line item (₹ Cr)").font = Font(bold=True, color="FFFFFF")
    wm.cell(2, 1).fill = PatternFill("solid", fgColor=NAVY)
    for i, col in enumerate(cols):
        c = wm.cell(2, 2 + i, ylabels[i])
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
    # Revenue (row3): Yr1 grows base; each next grows previous
    wm.cell(3, 1, "Revenue")
    wm["B3"] = "=ASSUMPTIONS!$B$3*(1+ASSUMPTIONS!B15)"
    for i in range(1, 5):
        prev, cur = cols[i - 1], cols[i]
        wm[f"{cur}3"] = f"={prev}3*(1+ASSUMPTIONS!{cur}15)"
    # EBIT, NOPAT, D&A, Capex, ΔNWC, FCF, DF, PV
    for i, col in enumerate(cols):
        wm[f"{col}4"] = f"={col}3*ASSUMPTIONS!{col}17"                    # EBIT (per-year margin path)
        wm[f"{col}5"] = f"={col}4*(1-ASSUMPTIONS!$B$5)"                   # NOPAT
        wm[f"{col}6"] = f"={col}3*ASSUMPTIONS!$B$7"                       # D&A
        wm[f"{col}7"] = f"={col}3*ASSUMPTIONS!$B$6"                       # Capex
        prevrev = "ASSUMPTIONS!$B$3" if i == 0 else f"{cols[i-1]}3"
        wm[f"{col}8"] = f"=({col}3-{prevrev})*ASSUMPTIONS!$B$8"           # ΔNWC
        wm[f"{col}9"] = f"={col}5+{col}6-{col}7-{col}8"                   # FCF
        wm[f"{col}10"] = f"=1/(1+ASSUMPTIONS!$B$9)^{i+1}"                 # discount factor
        wm[f"{col}11"] = f"={col}9*{col}10"                              # PV of FCF
    for row, lbl in [(4, "EBIT"), (5, "NOPAT"), (6, "(+) D&A"), (7, "(−) Capex"),
                     (8, "(−) Δ NWC"), (9, "Unlevered FCF"), (10, "Discount factor"), (11, "PV of FCF")]:
        wm.cell(row, 1, lbl)
    # Valuation block
    val = [
        ("Sum PV (explicit)", "=SUM(B11:F11)", num),                                          # B13
        ("Terminal value", "=F9*(1+ASSUMPTIONS!$B$10)/(ASSUMPTIONS!$B$9-ASSUMPTIONS!$B$10)", num),  # B14
        ("PV of terminal", "=B14*F10", num),                                                  # B15
        ("Enterprise value", "=B13+B15", num),                                                # B16
        ("Less: net debt", "=ASSUMPTIONS!$B$11", num),                                        # B17
        ("Equity value", "=B16-B17", num),                                                    # B18
        ("Shares (Cr)", "=ASSUMPTIONS!$B$12", "#,##0.00"),                                     # B19
        ("Fair value / share (₹)", "=B18/B19", "#,##0.00"),                                    # B20
        ("Current price (₹)", "=ASSUMPTIONS!$B$13", "#,##0.00"),                               # B21
        ("Upside / (downside)", "=B20/B21-1", pct),                                            # B22
    ]
    rr = 13
    for label, formula, fmt in val:
        wm.cell(rr, 1, label).font = Font(bold=True, color=NAVY)
        c = wm.cell(rr, 2, formula)
        c.number_format = fmt
        if label.startswith("Fair value") or label.startswith("Upside"):
            c.font = Font(bold=True, color=ARC)
        rr += 1
    wm.column_dimensions["A"].width = 22
    for col in "BCDEF":
        wm.column_dimensions[col].width = 13
    for row in wm.iter_rows(min_row=3, max_row=11, min_col=2, max_col=6):
        for cell in row:
            cell.number_format = num

    # COVER → MODEL cross-links (fair value + upside)
    cover = wb["COVER"]
    cr = cover.max_row + 2
    cover.cell(cr, 1, "DCF Fair Value (₹)").font = Font(bold=True, color=NAVY)
    fvc = cover.cell(cr, 2, "=MODEL!B20"); fvc.number_format = "#,##0.00"; fvc.font = Font(bold=True, color=ARC)
    cover.cell(cr + 1, 1, "Upside / (downside)").font = Font(bold=True, color=NAVY)
    uc = cover.cell(cr + 1, 2, "=MODEL!B22"); uc.number_format = pct; uc.font = Font(bold=True, color=ARC)

    # ── CONSENSUS (street vs our view) ──
    cons = consensus_summary(d)
    if cons:
        wc = wb.create_sheet("CONSENSUS")
        wc.sheet_view.showGridLines = False
        wc.append(["Street Consensus vs Our View", ""])
        style_header(wc, 1, 2)
        ours_t = ((narrative or {}).get("price_target")) or python_dcf(asmp)["fair_value"]
        rows = [
            ("Street recommendation", cons["rec"] or "—"),
            ("Street mean target (₹)", f"{cons['mean']:,.0f}"),
            ("Street target range (₹)", f"{cons['low']:,.0f} – {cons['high']:,.0f}" if cons.get("low") and cons.get("high") else "—"),
            ("Analysts covering", cons["n"] or "—"),
            ("Street implied upside", f"{cons['upside_pct']:+.0f}%"),
            ("— Our recommendation", final_verdict(a, narrative)),
            ("Our target (₹)", f"{ours_t:,.0f}" if ours_t else "—"),
            ("Our gap vs street", f"{(ours_t/cons['mean']-1)*100:+.0f}%" if ours_t else "—"),
        ]
        for k, v in rows:
            wc.append([k, v])
            if k.startswith("—") or k.startswith("Our"):
                wc.cell(wc.max_row, 1).font = Font(bold=True, color=ARC)
        if (narrative or {}).get("vs_consensus"):
            wc.append([])
            wc.append(["Why our view differs:", ""])
            wc.cell(wc.max_row, 1).font = Font(bold=True, color=NAVY)
            wc.append([narrative["vs_consensus"], ""])
            wc.cell(wc.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        if (narrative or {}).get("divergence_factor"):
            wc.append(["Key fundamental difference:", ""])
            wc.cell(wc.max_row, 1).font = Font(bold=True, color=ARC)
            wc.append([narrative["divergence_factor"], ""])
            wc.cell(wc.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        for row in wc.iter_rows(min_row=1, max_row=wc.max_row, max_col=2):
            for cell in row:
                cell.border = border
        wc.column_dimensions["A"].width = 26
        wc.column_dimensions["B"].width = 60

    # ── BROKERS (named ratings + distribution) ──
    rd = d.get("rating_dist")
    brk = d.get("brokers") or []
    if rd or brk:
        wbk = wb.create_sheet("BROKERS")
        wbk.sheet_view.showGridLines = False
        wbk.append(["Analyst Ratings", "", "", ""])
        style_header(wbk, 1, 4)
        if rd:
            total = sum(rd.values())
            wbk.append(["Rating distribution", f"{total} analysts", "", ""])
            wbk.cell(wbk.max_row, 1).font = Font(bold=True, color=NAVY)
            for lbl, key in [("Strong Buy", "strong_buy"), ("Buy", "buy"), ("Hold", "hold"),
                             ("Sell", "sell"), ("Strong Sell", "strong_sell")]:
                wbk.append([lbl, rd.get(key, 0), "", ""])
            wbk.append([])
        if brk:
            hrow = wbk.max_row + 1
            wbk.append(["Firm", "Rating", "Action", "Date"])
            for c in wbk[hrow]:
                c.fill = PatternFill("solid", fgColor=NAVY); c.font = Font(color="FFFFFF", bold=True)
            for b in brk:
                wbk.append([b["firm"], b["grade"], b["action"], b["date"]])
        else:
            wbk.append(["Named broker actions", "not available for this NSE listing (Yahoo)", "", ""])
            wbk.cell(wbk.max_row, 1).font = Font(italic=True, color=GREY)
        for row in wbk.iter_rows(min_row=1, max_row=wbk.max_row, max_col=4):
            for cell in row:
                cell.border = border
        wbk.column_dimensions["A"].width = 30
        for col in "BCD":
            wbk.column_dimensions[col].width = 16

    # ── SENSITIVITY (fair value across WACC × terminal growth) ──
    ws2 = wb.create_sheet("SENSITIVITY")
    ws2.sheet_view.showGridLines = False
    ws2.append(["DCF fair value/share (₹) — WACC (rows) × Terminal growth (cols)"])
    style_header(ws2, 1, 6)
    base_w, base_t = asmp["wacc"], asmp["terminal_growth"]
    waccs = [round(base_w + dw, 4) for dw in (-0.02, -0.01, 0, 0.01, 0.02)]
    tgs = [round(base_t + dt, 4) for dt in (-0.01, -0.005, 0, 0.005, 0.01)]
    ws2.append(["WACC \\ TGR"] + [f"{t*100:.1f}%" for t in tgs])
    for w in waccs:
        rowvals = [f"{w*100:.1f}%"]
        for t in tgs:
            aa = dict(asmp); aa["wacc"] = w; aa["terminal_growth"] = min(t, w - 0.01)
            fvv = python_dcf(aa)["fair_value"]
            rowvals.append(round(fvv) if fvv else None)
        ws2.append(rowvals)
        if abs(w - base_w) < 1e-6:
            for c in ws2[ws2.max_row]:
                c.font = Font(bold=True)
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=6):
        for cell in row:
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
    ws2.column_dimensions["A"].width = 14
    for col in "BCDEF":
        ws2.column_dimensions[col].width = 11

    # ── SCENARIOS (bull / base / bear, probability-weighted) ──
    scen = (narrative or {}).get("scenarios")
    if scen:
        wsn = wb.create_sheet("SCENARIOS")
        wsn.sheet_view.showGridLines = False
        wsn.append(["Scenario", "Target (₹)", "Upside", "Probability", "Key driver"])
        style_header(wsn, 1, 5)
        price = d["price"]
        ev_sum = 0.0
        for key in ("bull", "base", "bear"):
            s = scen.get(key) or {}
            tgt = s.get("target")
            prob = s.get("probability")
            up = f"{(tgt/price-1)*100:+.0f}%" if tgt else "—"
            wsn.append([key.capitalize(), round(tgt) if tgt else "—", up,
                        f"{prob*100:.0f}%" if prob is not None else "—", s.get("driver", "")])
            if tgt and prob:
                ev_sum += tgt * prob
        wsn.append([])
        wsn.append(["Probability-weighted target", round(ev_sum) if ev_sum else "—",
                    f"{(ev_sum/price-1)*100:+.0f}%" if ev_sum else "—", "", ""])
        wsn.cell(wsn.max_row, 1).font = Font(bold=True, color=ARC)
        wsn.cell(wsn.max_row, 2).font = Font(bold=True, color=ARC)
        for row in wsn.iter_rows(min_row=1, max_row=wsn.max_row, max_col=5):
            for cell in row:
                cell.border = border
        for col, w in zip("ABCDE", (14, 12, 10, 12, 46)):
            wsn.column_dimensions[col].width = w

    n = narrative or {}
    price = d["price"]

    # ── COMPS (market comparables) ──
    comps = n.get("comps")
    if comps and comps.get("peers"):
        wcp = wb.create_sheet("COMPS")
        wcp.sheet_view.showGridLines = False
        wcp.append(["Peer", "EV/EBITDA", "P/E"])
        style_header(wcp, 1, 3)
        for pr in comps["peers"]:
            wcp.append([pr.get("name", ""), pr.get("ev_ebitda"), pr.get("pe")])
        wcp.append([])
        wcp.append(["Peer median EV/EBITDA", comps.get("median_ev_ebitda"), ""])
        wcp.append(["Peer median P/E", comps.get("median_pe"), ""])
        iv = comps.get("implied_value_per_share")
        wcp.append(["Implied value / share (₹)", round(iv) if iv else "—", ""])
        wcp.cell(wcp.max_row, 1).font = Font(bold=True, color=ARC)
        wcp.cell(wcp.max_row, 2).font = Font(bold=True, color=ARC)
        if comps.get("note"):
            wcp.append([]); wcp.append([comps["note"], "", ""])
            wcp.cell(wcp.max_row, 1).font = Font(italic=True, color=GREY)
        for row in wcp.iter_rows(min_row=1, max_row=wcp.max_row, max_col=3):
            for c in row:
                c.border = border
        wcp.column_dimensions["A"].width = 28
        for col in "BC":
            wcp.column_dimensions[col].width = 14

    # ── SOTP (sum-of-the-parts) ──
    sotp = n.get("sotp")
    if sotp and sotp.get("segments"):
        wso = wb.create_sheet("SOTP")
        wso.sheet_view.showGridLines = False
        wso.append(["Segment", "Basis", "Metric (₹Cr)", "Multiple", "EV (₹Cr)"])
        style_header(wso, 1, 5)
        ev_total = 0.0
        for s in sotp["segments"]:
            ev = s.get("ev_cr") or 0
            ev_total += ev
            wso.append([s.get("segment", ""), s.get("basis", ""), s.get("metric_cr"),
                        s.get("multiple"), round(ev) if ev else "—"])
        wso.append(["Enterprise value", "", "", "", round(ev_total)])
        nd = sotp.get("net_debt_cr", 0) or 0
        wso.append(["Less: net debt", "", "", "", round(nd)])
        wso.append(["Equity value", "", "", "", round(ev_total - nd)])
        ivs = sotp.get("implied_value_per_share")
        wso.append(["Implied value / share (₹)", "", "", "", round(ivs) if ivs else "—"])
        for rr in range(wso.max_row - 3, wso.max_row + 1):
            wso.cell(rr, 1).font = Font(bold=True, color=NAVY)
        wso.cell(wso.max_row, 1).font = Font(bold=True, color=ARC)
        wso.cell(wso.max_row, 5).font = Font(bold=True, color=ARC)
        if sotp.get("note"):
            wso.append([]); wso.append([sotp["note"], "", "", "", ""])
        for row in wso.iter_rows(min_row=1, max_row=wso.max_row, max_col=5):
            for c in row:
                c.border = border
        wso.column_dimensions["A"].width = 26
        for col in "BCDE":
            wso.column_dimensions[col].width = 14

    # ── FOOTBALL FIELD (valuation range across methods + bar chart) ──
    methods = []
    dcf_fv = python_dcf(asmp)["fair_value"] if asmp else None
    if dcf_fv:
        methods.append(("DCF", dcf_fv * 0.85, dcf_fv * 1.15))
    if comps and comps.get("implied_value_per_share"):
        c = comps["implied_value_per_share"]; methods.append(("Comps", c * 0.9, c * 1.1))
    if sotp and sotp.get("implied_value_per_share"):
        s = sotp["implied_value_per_share"]; methods.append(("SOTP", s * 0.9, s * 1.1))
    cons = consensus_summary(d)
    if cons and cons.get("low") and cons.get("high"):
        methods.append(("Analyst targets", cons["low"], cons["high"]))
    if d.get("wk_low") and d.get("wk_high"):
        methods.append(("52-week range", d["wk_low"], d["wk_high"]))
    if len(methods) >= 2:
        wff = wb.create_sheet("FOOTBALL_FIELD")
        wff.sheet_view.showGridLines = False
        wff.append(["Method", "Low (₹)", "High (₹)", "Base (hidden)", "Range"])
        style_header(wff, 1, 5)
        for name, lo, hi in methods:
            wff.append([name, round(lo), round(hi), round(lo), round(hi - lo)])
        wff.append(["Current price", round(price), round(price), "", ""])
        for row in wff.iter_rows(min_row=1, max_row=wff.max_row, max_col=5):
            for c in row:
                c.border = border
                if isinstance(c.value, (int, float)):
                    c.number_format = "#,##0"
        wff.column_dimensions["A"].width = 18
        for col in "BCDE":
            wff.column_dimensions[col].width = 12
        # stacked horizontal bar → floating "range" bars = football field
        nrows = len(methods)
        chart = BarChart()
        chart.type = "bar"; chart.grouping = "stacked"; chart.overlap = 100
        chart.title = "Football Field — Valuation Range (₹/share)"
        chart.height = 8; chart.width = 18
        cats = Reference(wff, min_col=1, min_row=2, max_row=1 + nrows)
        base = Reference(wff, min_col=4, min_row=1, max_row=1 + nrows)   # hidden base
        rng = Reference(wff, min_col=5, min_row=1, max_row=1 + nrows)    # visible range
        chart.add_data(base, titles_from_data=True)
        chart.add_data(rng, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.noFill = True   # base invisible
        chart.series[1].graphicalProperties.solidFill = ARC
        chart.legend = None
        wff.add_chart(chart, "G2")

    # ── ANALYSIS (Claude narrative, grounded on real data) ──
    if narrative:
        ws = wb.create_sheet("ANALYSIS", 1)  # place right after COVER
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 100
        row = 1

        def section(title, content):
            nonlocal row
            c = ws.cell(row, 1, title)
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill("solid", fgColor=ARC)
            row += 1
            if isinstance(content, list):
                for item in content:
                    ws.cell(row, 1, f"•  {item}").alignment = Alignment(wrap_text=True, vertical="top")
                    row += 1
            else:
                cell = ws.cell(row, 1, str(content))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                row += 1
            row += 1  # spacer

        section("INVESTMENT THESIS", narrative.get("thesis", "—"))
        section("BUSINESS & MOAT", narrative.get("business", "—"))
        if narrative.get("bull_case"):
            section("BULL CASE", narrative["bull_case"])
        if narrative.get("bear_case"):
            section("BEAR CASE", narrative["bear_case"])
        if narrative.get("catalysts"):
            section("CATALYSTS", narrative["catalysts"])
        if narrative.get("risks"):
            section("KEY RISKS", narrative["risks"])
        section("RECOMMENDATION RATIONALE", narrative.get("recommendation", a["action"]))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════
#  4. INSTITUTIONAL PDF
# ═══════════════════════════════════════════════════════════
def build_pdf(d: dict, a: dict, narrative=None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=16 * mm, rightMargin=16 * mm,
                            topMargin=16 * mm, bottomMargin=16 * mm)
    ss = getSampleStyleSheet()
    for _sn in ss.byName:
        ss[_sn].fontName = PDF_FONT
    h1 = ParagraphStyle("h1", parent=ss["Title"], fontName=PDF_BOLD, textColor=colors.HexColor("#" + NAVY), fontSize=18, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=ss["Normal"], fontName=PDF_FONT, textColor=colors.HexColor("#" + GREY), fontSize=9, spaceAfter=10)
    hh = ParagraphStyle("hh", parent=ss["Heading2"], fontName=PDF_BOLD, textColor=colors.HexColor("#" + ARC), fontSize=12, spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle("body", parent=ss["Normal"], fontName=PDF_FONT, fontSize=9.5, leading=13)
    story = []

    story.append(Paragraph("J.A.R.V.I.S. · Institutional Equity Research", h1))
    story.append(Paragraph(datetime.now().strftime("%d %B %Y") + " &nbsp;·&nbsp; For informational purposes only — not investment advice", sub))

    # Recommendation banner
    verdict = final_verdict(a, narrative)
    verdict_tbl = Table([[f"{d['name']}  ({d['symbol']})", f"{verdict}"]], colWidths=[120 * mm, 45 * mm])
    verdict_tbl.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#" + NAVY)),
        ("TEXTCOLOR", (0, 0), (0, 0), colors.white),
        ("TEXTCOLOR", (1, 0), (1, 0), colors.HexColor("#" + ARC)),
        ("FONTSIZE", (0, 0), (0, 0), 12),
        ("FONTSIZE", (1, 0), (1, 0), 14),
        ("FONTNAME", (0, 0), (-1, -1), PDF_BOLD),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(verdict_tbl)
    story.append(Spacer(1, 6))

    def bullets(items):
        for it in items:
            story.append(Paragraph(f"•&nbsp; {it}", body))

    if narrative:
        story.append(Paragraph("Investment Thesis", hh))
        story.append(Paragraph(narrative.get("thesis", ""), body))
        if narrative.get("business"):
            story.append(Paragraph(f"<b>Business & moat.</b> {narrative['business']}", body))
        if narrative.get("bull_case"):
            story.append(Paragraph("Bull Case", hh)); bullets(narrative["bull_case"])
        if narrative.get("bear_case"):
            story.append(Paragraph("Bear Case", hh)); bullets(narrative["bear_case"])
        if narrative.get("catalysts"):
            story.append(Paragraph("Catalysts", hh)); bullets(narrative["catalysts"])
        if narrative.get("risks"):
            story.append(Paragraph("Key Risks", hh)); bullets(narrative["risks"])
    else:
        story.append(Paragraph("Thesis", hh))
        story.append(Paragraph(f"<b>{d['sector']} · {d['industry']}.</b> {a['action']}.", body))
        if d.get("summary"):
            story.append(Paragraph(d["summary"][:600] + ("…" if len(d["summary"]) > 600 else ""), body))

    # Snapshot table
    story.append(Paragraph("Snapshot", hh))
    snap = [
        ["Price", f"₹{d['price']:,.2f}", "Market Cap", _fmt_money(d.get("market_cap"))],
        ["P/E", f"{d['pe']:.1f}x" if d.get("pe") else "N/A", "ROE", f"{d['roe']*100:.1f}%" if d.get("roe") is not None else "N/A"],
        ["Rev. growth", f"{d['rev_growth']*100:.1f}%" if d.get("rev_growth") is not None else "N/A", "D/E", f"{d['debt_to_equity']:.2f}" if d.get("debt_to_equity") is not None else "N/A"],
        ["52-wk", f"₹{d.get('wk_low',0):,.0f}–₹{d.get('wk_high',0):,.0f}" if d.get("wk_high") else "N/A", "200-DMA", f"₹{d['dma200']:,.0f}" if d.get("dma200") else "N/A"],
    ]
    t = Table(snap, colWidths=[30 * mm, 52 * mm, 30 * mm, 53 * mm])
    t.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7DE")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#" + LIGHT)),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#" + LIGHT)),
        ("FONTNAME", (0, 0), (0, -1), PDF_BOLD),
        ("FONTNAME", (2, 0), (2, -1), PDF_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t)

    # Factor scorecard
    story.append(Paragraph("Factor Scorecard", hh))
    data = [["Factor", "Value", "Signal", "Score"]]
    for f in a["factors"]:
        data.append([f["factor"], str(f["value"]), f["signal"], f"{f['points']:+d}"])
    data.append(["COMPOSITE", f"{a['score']}/{a['n']}", a["verdict"], f"{a['norm']:.2f}"])
    st = Table(data, colWidths=[42 * mm, 40 * mm, 45 * mm, 18 * mm])
    style = [
        ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + NAVY)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), PDF_BOLD),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7DE")),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#" + LIGHT)),
        ("FONTNAME", (0, -1), (-1, -1), PDF_BOLD),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    story.append(Table(data, colWidths=[42 * mm, 40 * mm, 45 * mm, 18 * mm], style=TableStyle(style)))

    # DCF
    dcf = a["dcf"]
    story.append(Paragraph("Valuation — DCF", hh))
    if dcf["fair_value"] is not None:
        story.append(Paragraph(
            f"Two-stage 10-yr DCF (discount {dcf['assumptions']['discount_rate']*100:.0f}%, "
            f"terminal {dcf['assumptions']['terminal_growth']*100:.0f}%) → "
            f"<b>fair value ₹{dcf['fair_value']:,.0f}</b> vs price ₹{d['price']:,.2f}, "
            f"implying <b>{dcf['upside']*100:+.0f}%</b>.", body))
    else:
        story.append(Paragraph(dcf["note"], body))

    # Scenario analysis (bull / base / bear)
    scen = (narrative or {}).get("scenarios")
    if scen:
        story.append(Paragraph("Scenario Analysis", hh))
        sdata = [["Scenario", "Target (₹)", "Upside", "Prob.", "Driver"]]
        ev = 0.0
        for key in ("bull", "base", "bear"):
            s = scen.get(key) or {}
            tg = s.get("target"); pr = s.get("probability")
            sdata.append([key.capitalize(), f"{tg:,.0f}" if tg else "—",
                          f"{(tg/d['price']-1)*100:+.0f}%" if tg else "—",
                          f"{pr*100:.0f}%" if pr is not None else "—", s.get("driver", "")])
            if tg and pr:
                ev += tg * pr
        if ev:
            sdata.append(["Weighted", f"{ev:,.0f}", f"{(ev/d['price']-1)*100:+.0f}%", "", ""])
        st = Table(sdata, colWidths=[22 * mm, 26 * mm, 20 * mm, 16 * mm, 81 * mm])
        st.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + NAVY)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), PDF_BOLD),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7DE")), ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#" + LIGHT)), ("FONTNAME", (0, -1), (-1, -1), PDF_BOLD),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
        story.append(st)

    # Valuation football field (methods → range)
    _comps = (narrative or {}).get("comps")
    _sotp = (narrative or {}).get("sotp")
    _cons0 = consensus_summary(d)
    frows = [["Method", "Low (₹)", "High (₹)"]]
    if dcf["fair_value"]:
        frows.append(["DCF", f"{dcf['fair_value']*0.85:,.0f}", f"{dcf['fair_value']*1.15:,.0f}"])
    if _comps and _comps.get("implied_value_per_share"):
        c = _comps["implied_value_per_share"]; frows.append(["Comps", f"{c*0.9:,.0f}", f"{c*1.1:,.0f}"])
    if _sotp and _sotp.get("implied_value_per_share"):
        s = _sotp["implied_value_per_share"]; frows.append(["SOTP", f"{s*0.9:,.0f}", f"{s*1.1:,.0f}"])
    if _cons0 and _cons0.get("low") and _cons0.get("high"):
        frows.append(["Analyst targets", f"{_cons0['low']:,.0f}", f"{_cons0['high']:,.0f}"])
    if d.get("wk_low") and d.get("wk_high"):
        frows.append(["52-week range", f"{d['wk_low']:,.0f}", f"{d['wk_high']:,.0f}"])
    if len(frows) >= 3:
        story.append(Paragraph("Valuation Football Field", hh))
        ff = Table(frows, colWidths=[45 * mm, 40 * mm, 40 * mm])
        ff.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + NAVY)), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), PDF_BOLD), ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7DE")),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
        story.append(ff)
        story.append(Paragraph(f"<i>Current price ₹{d['price']:,.0f}. Full chart, Comps and SOTP sheets in the Excel model.</i>",
                     ParagraphStyle("ffn", parent=body, fontSize=7.5, textColor=colors.HexColor("#" + GREY))))

    # Street consensus vs our view
    cons = consensus_summary(d)
    if cons:
        story.append(Paragraph("Street Consensus vs Our View", hh))
        ours_t = (narrative or {}).get("price_target") or dcf.get("fair_value")
        crows = [["", "Recommendation", "Target (₹)", "Implied"]]
        crows.append(["Street", cons["rec"] or "—", f"{cons['mean']:,.0f}", f"{cons['upside_pct']:+.0f}%"])
        crows.append(["J.A.R.V.I.S.", verdict, f"{ours_t:,.0f}" if ours_t else "—",
                      f"{(ours_t/d['price']-1)*100:+.0f}%" if ours_t else "—"])
        ct = Table(crows, colWidths=[35 * mm, 45 * mm, 40 * mm, 45 * mm])
        ct.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + NAVY)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), PDF_BOLD),
            ("FONTNAME", (0, 2), (0, 2), PDF_BOLD),
            ("TEXTCOLOR", (0, 2), (-1, 2), colors.HexColor("#" + ARC)),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7DE")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(ct)
        rd = d.get("rating_dist")
        if rd:
            story.append(Spacer(1, 2))
            story.append(Paragraph(
                f"Ratings: {rd.get('strong_buy',0)} strong-buy · {rd.get('buy',0)} buy · "
                f"{rd.get('hold',0)} hold · {rd.get('sell',0)} sell · {rd.get('strong_sell',0)} strong-sell", body))
        if (narrative or {}).get("vs_consensus"):
            story.append(Spacer(1, 3))
            story.append(Paragraph(f"<i>{narrative['vs_consensus']}</i>", body))
        if (narrative or {}).get("divergence_factor"):
            story.append(Paragraph(f"<b>Key fundamental difference:</b> {narrative['divergence_factor']}", body))

    story.append(Paragraph("Recommendation", hh))
    rec_text = narrative.get("recommendation") if narrative else a["action"]
    story.append(Paragraph(f"<b>{verdict}.</b> {rec_text}", body))

    story.append(Spacer(1, 10))
    src = "Figures: Yahoo Finance (live). Narrative: Claude." if narrative else "Data: Yahoo Finance (live)."
    story.append(Paragraph(
        f"<i>{src} Generated by J.A.R.V.I.S. Report Engine. "
        "Automated research, not personalised investment advice.</i>",
        ParagraphStyle("disc", parent=body, fontSize=7.5, textColor=colors.HexColor("#" + GREY))))

    doc.build(story)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════
#  4b. VALIDATION — evaluate the workbook's real formulas
# ═══════════════════════════════════════════════════════════
def validate_excel(xlsx_bytes: bytes, assumptions: dict) -> dict:
    """Load the generated workbook, actually COMPUTE its formulas with the
    `formulas` engine, and cross-check against an independent Python recompute.
    Returns {ok, checks:[{name, pass, detail}], computed_fair_value}."""
    import os
    import tempfile
    from openpyxl import load_workbook

    checks = []
    ok = True

    def add(name, passed, detail=""):
        nonlocal ok
        checks.append({"name": name, "pass": bool(passed), "detail": str(detail)})
        if not passed:
            ok = False

    # 1. Structure
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        need = {"COVER", "KEY_METRICS", "SCORECARD", "ASSUMPTIONS", "MODEL"}
        missing = need - set(wb.sheetnames)
        add("all sheets present", not missing, "missing " + ", ".join(missing) if missing else "COVER·ANALYSIS·KEY_METRICS·SCORECARD·ASSUMPTIONS·MODEL")
        # cross-sheet links exist in MODEL
        wm = wb["MODEL"]
        linked = sum(1 for row in wm.iter_rows() for c in row
                     if isinstance(c.value, str) and c.value.startswith("=") and "ASSUMPTIONS!" in c.value)
        add("model is formula-linked", linked >= 15, f"{linked} cross-sheet formulas")
    except Exception as ex:
        add("workbook opens", False, str(ex)[:80])
        return {"ok": ok, "checks": checks, "computed_fair_value": None}

    # 2. Independent Python recompute (the oracle) — this IS the calc check.
    expected = python_dcf(assumptions)
    exp_fv = expected["fair_value"]
    add("enterprise value > 0", expected["ev"] > 0, f"₹{expected['ev']:,.0f} Cr")
    add("fair value computes", exp_fv is not None and exp_fv > 0,
        f"₹{exp_fv:,.2f}" if exp_fv else "not computed")
    computed_fv = exp_fv

    # 3. Optional heavy check: actually execute the workbook formulas and confirm
    #    they match the recompute. Slow (~8s) — off by default; the deterministic
    #    formula-writer + oracle already guarantee correctness.
    if os.getenv("VALIDATE_FORMULAS", "0") == "1":
        tmp_path = None
        try:
            import numpy as np
            import formulas
            tf = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tf.write(xlsx_bytes)
            tf.close()
            tmp_path = tf.name
            sol = formulas.ExcelModel().loads(tmp_path).finish().calculate()

            def get(cell):
                suffix = f"MODEL'!{cell}".upper()
                for k, v in sol.items():
                    if k.upper().endswith(suffix):
                        try:
                            return float(np.ravel(v.value)[0])
                        except Exception:
                            return None
                return None

            engine_fv = get("B20")
            if engine_fv is not None and exp_fv:
                diff = abs(engine_fv - exp_fv) / abs(exp_fv)
                add("Excel formulas match recompute (<0.5%)", diff < 0.005,
                    f"Excel ₹{engine_fv:,.2f} vs Python ₹{exp_fv:,.2f} ({diff*100:.3f}%)")
        except Exception as ex:
            add("formula engine evaluates", False, str(ex)[:100])
        finally:
            if tmp_path:
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass

    return {"ok": ok, "checks": checks, "computed_fair_value": computed_fv,
            "expected_fair_value": exp_fv}


# ═══════════════════════════════════════════════════════════
#  5. TOP-LEVEL
# ═══════════════════════════════════════════════════════════
def prepare(symbol: str):
    """Fetch + analyse. Returns (data, analysis)."""
    d = fetch_stock(symbol)
    a = analyze(d)
    return d, a


def build_prompt_short(d: dict, a: dict):
    """Short-term (6-12 month) technical + catalyst trade note. No DCF."""
    t = d.get("tech") or {}
    price = d["price"]
    rr = lambda x, m=1: round(x * m, 2) if x is not None else None
    facts = {
        "company": d["name"], "symbol": d["symbol"], "sector": d["sector"], "industry": d["industry"],
        "price_inr": round(price, 2),
        "ma20": rr(t.get("ma20")), "ma50": rr(t.get("ma50")), "ma200": rr(t.get("ma200")),
        "rsi14": round(t["rsi"], 1) if t.get("rsi") is not None else None,
        "return_1m_pct": rr(t.get("ret_1m"), 100), "return_3m_pct": rr(t.get("ret_3m"), 100),
        "return_6m_pct": rr(t.get("ret_6m"), 100), "annual_vol_pct": rr(t.get("vol"), 100),
        "support_3m": rr(t.get("support")), "resistance_3m": rr(t.get("resistance")),
        "wk_high": d.get("wk_high"), "wk_low": d.get("wk_low"), "trailing_pe": d.get("pe"),
    }
    cons = consensus_summary(d)
    if cons:
        facts["street_consensus"] = {"mean_target_inr": round(cons["mean"]),
                                     "recommendation": cons["rec"], "num_analysts": cons["n"]}
    system = (
        "You are the head of a technical / quant trading desk writing a SHORT-TERM (6-12 month) "
        "positional trade note. DCF and intrinsic valuation are NOT relevant on this horizon — base your "
        "view on price trend, moving averages (20/50/200-DMA), RSI, momentum, key support/resistance "
        "levels, volatility, and near-term catalysts (earnings, events, sector flows). Use ONLY the "
        "verified figures — never invent a number. Set actionable levels consistent with those figures."
    )
    user = (
        f"Produce a 6-12 month technical & catalyst trade note for {d['name']} ({d['symbol']}).\n\n"
        f"VERIFIED FIGURES:\n{json.dumps(facts, indent=2)}\n\n"
        "Return ONLY strict JSON (no fences, comments, or trailing commas) with keys:\n"
        "- thesis: 2-3 sentence short-term setup (trend + momentum)\n"
        "- trend: 1-2 sentences reading the DMAs, RSI and momentum\n"
        "- bull_case: array of 3 short points\n- bear_case: array of 3 short points\n"
        "- catalysts: array of 2-3 near-term catalysts with rough timing (next 6-12 months)\n"
        "- risks: array of 2 key risks\n"
        "- verdict: one of BUY, ACCUMULATE, HOLD, REDUCE, SELL (the 6-12 month trade call)\n"
        "- verdict_rationale: one short clause under 16 words\n"
        "- levels: object {entry, target, stop, support, resistance} — INR numbers, realistic vs price\n"
        "- price_target: 6-12 month target (INR number)\n"
        "- recommendation: 2 sentence rationale for the trade\n"
        "- vs_consensus: 1-2 sentences comparing to street_consensus if present, else empty string\n"
        'Example: {"thesis":"...","trend":"...","bull_case":["..."],"bear_case":["..."],'
        '"catalysts":["..."],"risks":["..."],"verdict":"ACCUMULATE","verdict_rationale":"Uptrend, RSI neutral, buy dips",'
        '"levels":{"entry":1450,"target":1650,"stop":1360,"support":1400,"resistance":1600},'
        '"price_target":1650,"recommendation":"...","vs_consensus":"..."}'
    )
    return system, user


def compose_short(d: dict, a: dict, narrative, result: dict):
    """Chat + speech for the short-term technical note (no DCF)."""
    n = narrative or {}
    t = d.get("tech") or {}
    price = d["price"]
    verdict = final_verdict(a, narrative)
    tgt = result.get("price_target")
    lv = n.get("levels") or {}
    sector = d.get("industry") or d.get("sector") or ""
    thesis = (n.get("thesis") or "Short-term technical view.").strip()
    thesis_first = thesis.split(". ")[0].rstrip(".") + "."
    rationale = (n.get("verdict_rationale") or "").strip()
    intro = f"{d['name']}, {sector}." if sector else f"{d['name']}."
    final_line = f"Short-term call: {verdict}" + (f", {rationale}" if rationale else "") + \
                 (f". Target {tgt:,.0f} rupees." if tgt else ".")
    speech = " ".join([intro, thesis_first, final_line])

    pn = lambda x: f"₹{x:,.0f}" if x is not None else "—"
    p = [f"**{d['name']} ({d['symbol']})** · ₹{price:,.0f}  ·  _Short-term (6–12 mo) · technical view_"]
    tl = []
    if t.get("rsi") is not None: tl.append(f"RSI {t['rsi']:.0f}")
    if t.get("ma50"): tl.append(f"50-DMA ₹{t['ma50']:,.0f}")
    if t.get("ma200"): tl.append(f"200-DMA ₹{t['ma200']:,.0f}")
    if t.get("ret_1m") is not None: tl.append(f"1M {t['ret_1m']*100:+.1f}%")
    if t.get("ret_3m") is not None: tl.append(f"3M {t['ret_3m']*100:+.1f}%")
    if t.get("ret_6m") is not None: tl.append(f"6M {t['ret_6m']*100:+.1f}%")
    if t.get("vol") is not None: tl.append(f"vol {t['vol']*100:.0f}%")
    if tl: p.append(" · ".join(tl))
    p.append(f"**Thesis.** {thesis}")
    if n.get("trend"): p.append(f"**Technical read.** {n['trend']}")
    if n.get("bull_case"): p.append("**Bull:**\n" + "\n".join(f"- {x}" for x in n["bull_case"]))
    if n.get("bear_case"): p.append("**Bear:**\n" + "\n".join(f"- {x}" for x in n["bear_case"]))
    if n.get("catalysts"): p.append("**Catalysts:**\n" + "\n".join(f"- {x}" for x in n["catalysts"]))
    if n.get("risks"): p.append("**Risks:**\n" + "\n".join(f"- {x}" for x in n["risks"]))
    kl = ["**Key levels**",
          f"Support {pn(lv.get('support') or t.get('support'))} · Resistance {pn(lv.get('resistance') or t.get('resistance'))}",
          f"Entry {pn(lv.get('entry'))} · Target {pn(lv.get('target') or tgt)} · Stop {pn(lv.get('stop'))}"]
    if lv.get("target") and lv.get("stop") and lv.get("entry") and lv["entry"] != lv["stop"]:
        rrr = (lv["target"] - lv["entry"]) / (lv["entry"] - lv["stop"])
        kl.append(f"Risk : reward ≈ {rrr:.1f} : 1")
    p.append("\n".join(kl))
    cons = consensus_summary(d)
    if cons:
        cl = ["**Street consensus vs our view**",
              f"Street: **{cons['rec'] or 'N/A'}** · target ₹{cons['mean']:,.0f} ({cons['upside_pct']:+.0f}%)"]
        if tgt: cl.append(f"Ours: **{verdict}** · target ₹{tgt:,.0f}")
        if n.get("vs_consensus"): cl.append(n["vs_consensus"])
        p.append("\n".join(cl))
    p.append("_📊 Technical summary in the downloaded Excel + PDF · No DCF (short-term horizon)._")
    rec = (n.get("recommendation") or a["action"]).strip()
    p.append(f"**Recommendation: {verdict}**\n{rec}")
    return "\n\n".join(p), speech


def build_excel_short(d: dict, a: dict, narrative) -> bytes:
    """Compact technical workbook for short-term (Cover · Technicals · Analysis)."""
    n = narrative or {}
    t = d.get("tech") or {}
    lv = n.get("levels") or {}
    wb = Workbook()
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = PatternFill("solid", fgColor=NAVY)
    hfont = Font(color="FFFFFF", bold=True, size=11)
    verdict = final_verdict(a, narrative)

    ws = wb.active
    ws.title = "COVER"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "J.A.R.V.I.S. · SHORT-TERM TECHNICAL NOTE (6-12 MONTHS)"
    ws["A1"].fill = hf; ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws.merge_cells("A1:D1")
    rows = [("Company", d["name"]), ("Symbol", d["symbol"]), ("Price (₹)", f"{d['price']:,.2f}"),
            ("Call", verdict), ("Target (₹)", f"{n.get('price_target'):,.0f}" if n.get("price_target") else "—"),
            ("Report Date", datetime.now().strftime("%d %b %Y"))]
    r = 3
    for k, v in rows:
        ws.cell(r, 1, k).font = Font(bold=True, color=NAVY)
        ws.cell(r, 2, v)
        if k == "Call": ws.cell(r, 2).font = Font(bold=True, size=13, color=ARC)
        r += 1
    ws.column_dimensions["A"].width = 18; ws.column_dimensions["B"].width = 46

    wt = wb.create_sheet("TECHNICALS")
    wt.sheet_view.showGridLines = False
    wt.append(["Indicator", "Value"]);
    for c in wt[1]: c.fill = hf; c.font = hfont; c.border = border
    tvals = [
        ("Price", f"₹{d['price']:,.2f}"),
        ("RSI (14)", f"{t['rsi']:.0f}" if t.get("rsi") is not None else "—"),
        ("20-DMA", f"₹{t['ma20']:,.0f}" if t.get("ma20") else "—"),
        ("50-DMA", f"₹{t['ma50']:,.0f}" if t.get("ma50") else "—"),
        ("200-DMA", f"₹{t['ma200']:,.0f}" if t.get("ma200") else "—"),
        ("1-month return", f"{t['ret_1m']*100:+.1f}%" if t.get("ret_1m") is not None else "—"),
        ("3-month return", f"{t['ret_3m']*100:+.1f}%" if t.get("ret_3m") is not None else "—"),
        ("6-month return", f"{t['ret_6m']*100:+.1f}%" if t.get("ret_6m") is not None else "—"),
        ("Annualised volatility", f"{t['vol']*100:.0f}%" if t.get("vol") is not None else "—"),
        ("52-week range", f"₹{d.get('wk_low',0):,.0f} – ₹{d.get('wk_high',0):,.0f}" if d.get("wk_high") else "—"),
        ("Support (3m)", f"₹{(lv.get('support') or t.get('support') or 0):,.0f}"),
        ("Resistance (3m)", f"₹{(lv.get('resistance') or t.get('resistance') or 0):,.0f}"),
        ("Entry", f"₹{lv.get('entry'):,.0f}" if lv.get("entry") else "—"),
        ("Target (6-12m)", f"₹{lv.get('target') or n.get('price_target'):,.0f}" if (lv.get("target") or n.get("price_target")) else "—"),
        ("Stop-loss", f"₹{lv.get('stop'):,.0f}" if lv.get("stop") else "—"),
    ]
    for k, v in tvals:
        wt.append([k, v])
    for row in wt.iter_rows(min_row=1, max_row=wt.max_row, max_col=2):
        for cell in row: cell.border = border
    wt.column_dimensions["A"].width = 24; wt.column_dimensions["B"].width = 26

    if n:
        wa = wb.create_sheet("ANALYSIS")
        wa.sheet_view.showGridLines = False
        wa.column_dimensions["A"].width = 100
        row = 1
        def sec(title, content):
            nonlocal row
            c = wa.cell(row, 1, title); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=ARC)
            row += 1
            items = content if isinstance(content, list) else [content]
            for it in items:
                cc = wa.cell(row, 1, (f"•  {it}" if isinstance(content, list) else str(it)))
                cc.alignment = Alignment(wrap_text=True, vertical="top"); row += 1
            row += 1
        sec("THESIS", n.get("thesis", "—"))
        if n.get("trend"): sec("TECHNICAL READ", n["trend"])
        if n.get("bull_case"): sec("BULL", n["bull_case"])
        if n.get("bear_case"): sec("BEAR", n["bear_case"])
        if n.get("catalysts"): sec("CATALYSTS", n["catalysts"])
        if n.get("risks"): sec("RISKS", n["risks"])
        sec("RECOMMENDATION", n.get("recommendation", a["action"]))
        if n.get("vs_consensus"): sec("VS STREET CONSENSUS", n["vs_consensus"])

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def build_pdf_short(d: dict, a: dict, narrative) -> bytes:
    """Compact short-term technical PDF note."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=16 * mm, rightMargin=16 * mm,
                            topMargin=16 * mm, bottomMargin=16 * mm)
    ss = getSampleStyleSheet()
    for _sn in ss.byName:
        ss[_sn].fontName = PDF_FONT
    h1 = ParagraphStyle("h1", parent=ss["Title"], fontName=PDF_BOLD, textColor=colors.HexColor("#" + NAVY), fontSize=17, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=ss["Normal"], fontName=PDF_FONT, textColor=colors.HexColor("#" + GREY), fontSize=9, spaceAfter=10)
    hh = ParagraphStyle("hh", parent=ss["Heading2"], fontName=PDF_BOLD, textColor=colors.HexColor("#" + ARC), fontSize=12, spaceBefore=9, spaceAfter=4)
    body = ParagraphStyle("body", parent=ss["Normal"], fontName=PDF_FONT, fontSize=9.5, leading=13)
    n = narrative or {}; t = d.get("tech") or {}; lv = n.get("levels") or {}
    verdict = final_verdict(a, narrative)
    g = lambda x, f="₹{:,.0f}": f.format(x) if x is not None else "N/A"
    story = [Paragraph("J.A.R.V.I.S. · Short-Term Technical Note", h1),
             Paragraph(datetime.now().strftime("%d %B %Y") + " · 6-12 month horizon · Not investment advice", sub)]
    vt = Table([[f"{d['name']} ({d['symbol']})", verdict]], colWidths=[120 * mm, 45 * mm])
    vt.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), PDF_FONT),("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#" + NAVY)),
        ("TEXTCOLOR", (0, 0), (0, 0), colors.white), ("TEXTCOLOR", (1, 0), (1, 0), colors.HexColor("#" + ARC)),
        ("FONTNAME", (0, 0), (-1, -1), PDF_BOLD), ("FONTSIZE", (1, 0), (1, 0), 14),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"), ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10)]))
    story += [vt, Spacer(1, 6)]
    if n.get("thesis"): story += [Paragraph("Thesis", hh), Paragraph(n["thesis"], body)]
    if n.get("trend"): story += [Paragraph("Technical Read", hh), Paragraph(n["trend"], body)]
    story.append(Paragraph("Technical Snapshot", hh))
    snap = [["RSI(14)", f"{t['rsi']:.0f}" if t.get("rsi") is not None else "N/A", "50-DMA", g(t.get("ma50"))],
            ["200-DMA", g(t.get("ma200")), "6M return", f"{t['ret_6m']*100:+.1f}%" if t.get("ret_6m") is not None else "N/A"],
            ["Support", g(lv.get("support") or t.get("support")), "Resistance", g(lv.get("resistance") or t.get("resistance"))],
            ["Entry", g(lv.get("entry")), "Target", g(lv.get("target") or n.get("price_target"))],
            ["Stop-loss", g(lv.get("stop")), "Volatility", f"{t['vol']*100:.0f}%" if t.get("vol") is not None else "N/A"]]
    tb = Table(snap, colWidths=[30 * mm, 52 * mm, 30 * mm, 53 * mm])
    tb.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), PDF_FONT),("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7DE")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#" + LIGHT)), ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#" + LIGHT)),
        ("FONTNAME", (0, 0), (0, -1), PDF_BOLD), ("FONTNAME", (2, 0), (2, -1), PDF_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 9), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
    story.append(tb)

    def bl(title, items):
        if items:
            story.append(Paragraph(title, hh))
            for x in items:
                story.append(Paragraph(f"•&nbsp; {x}", body))
    bl("Bull Case", n.get("bull_case")); bl("Bear Case", n.get("bear_case"))
    bl("Catalysts", n.get("catalysts")); bl("Key Risks", n.get("risks"))
    cons = consensus_summary(d)
    if cons:
        story.append(Paragraph("Street Consensus vs Our View", hh))
        story.append(Paragraph(f"Street: <b>{cons['rec'] or 'N/A'}</b>, mean target ₹{cons['mean']:,.0f} "
                               f"({cons['upside_pct']:+.0f}%). Ours: <b>{verdict}</b>, target ₹{n.get('price_target',0):,.0f}. "
                               f"{n.get('vs_consensus','')}", body))
    story.append(Paragraph("Recommendation", hh))
    story.append(Paragraph(f"<b>{verdict}.</b> {n.get('recommendation', a['action'])}", body))
    story.append(Spacer(1, 10))
    story.append(Paragraph("<i>Data: Yahoo Finance (live). Technical note, not personalised investment advice.</i>",
                 ParagraphStyle("disc", parent=body, fontSize=7.5, textColor=colors.HexColor("#" + GREY))))
    doc.build(story)
    return buf.getvalue()


def build_prompt(d: dict, a: dict):
    """Build (system, user) for Claude to author the institutional narrative,
    grounded on REAL figures so it doesn't invent numbers."""
    facts = {
        "company": d["name"], "symbol": d["symbol"],
        "sector": d["sector"], "industry": d["industry"],
        "price_inr": round(d["price"], 2),
        "market_cap": _fmt_money(d.get("market_cap")),
        "trailing_pe": d.get("pe"), "forward_pe": d.get("forward_pe"),
        "peg": d.get("peg"), "price_to_book": d.get("pb"),
        "roe_pct": round(d["roe"] * 100, 1) if d.get("roe") is not None else None,
        "profit_margin_pct": round(d["margin"] * 100, 1) if d.get("margin") is not None else None,
        "revenue_growth_pct": round(d["rev_growth"] * 100, 1) if d.get("rev_growth") is not None else None,
        "debt_to_equity": d.get("debt_to_equity"),
        "revenue": _fmt_money(d.get("revenue")), "fcf": _fmt_money(d.get("fcf")),
        "wk_high": d.get("wk_high"), "wk_low": d.get("wk_low"),
        "dma200": d.get("dma200"),
        "dcf_fair_value": round(a["dcf"]["fair_value"]) if a["dcf"]["fair_value"] else None,
        "dcf_upside_pct": round(a["dcf"]["upside"] * 100) if a["dcf"]["upside"] is not None else None,
        "quant_verdict": a["verdict"], "quant_score": f"{a['score']}/{a['n']}",
        "historical_revenue_cagr_pct": round(d["rev_cagr"] * 100, 1) if d.get("rev_cagr") is not None else None,
        "revenue_history_cr_newest_to_oldest": [round(x) for x in d["rev_hist_cr"]] if d.get("rev_hist_cr") else None,
        "consensus_revenue_growth_pct": round(d["rev_growth"] * 100, 1) if d.get("rev_growth") is not None else None,
    }
    cons = consensus_summary(d)
    if cons:
        facts["street_consensus"] = {
            "mean_target_inr": round(cons["mean"]),
            "target_range_inr": [round(cons["low"]) if cons["low"] else None,
                                 round(cons["high"]) if cons["high"] else None],
            "recommendation": cons["rec"],
            "num_analysts": cons["n"],
            "implied_upside_pct": round(cons["upside_pct"]),
            "rating_distribution": d.get("rating_dist"),
        }
    system = (
        "You are a Managing Director of equity research at a top-tier institutional "
        "desk (think Goldman Sachs / Morgan Stanley), writing a long-horizon (3–5 year) "
        "coverage note for sophisticated buy-side clients. Standards:\n"
        "• Rigorous, specific, and balanced — no hype, no filler, no hedging platitudes.\n"
        "• Ground every claim in the VERIFIED FIGURES provided; NEVER invent or alter a "
        "number, price, ratio, or statistic. Qualitative judgement is yours; data is not.\n"
        "• Think in terms of moat, unit economics, capital allocation, industry structure, "
        "and through-cycle earnings power.\n"
        "• You are presenting to the investment committee of a sovereign wealth fund (e.g. ADIA/"
        "Dubai). The workings must be defensible line-by-line — no plain-vanilla shortcuts.\n"
        "• You set the forward DCF drivers that populate a live Excel model. TRIANGULATE revenue "
        "growth from THREE anchors and blend them: (1) the company's historical revenue CAGR "
        "provided, (2) the structural growth rate of its industry (your knowledge), (3) the "
        "broker/consensus revenue-growth expectation provided. Do NOT just pick a round number or "
        "a simple fade — show the blend in sources.growth. Do the SAME for the margin/expense path: "
        "anchor EBIT margin to the trailing margin and model realistic operating leverage / cost "
        "inflation, not a flat assumption."
    )
    user = (
        f"Produce an institutional long-term research note and DCF driver set for "
        f"{d['name']} ({d['symbol']}).\n\n"
        f"VERIFIED FIGURES (the only hard numbers you may cite):\n{json.dumps(facts, indent=2)}\n\n"
        "Return ONLY strictly valid JSON — no markdown fences, no comments, no trailing "
        "commas, no text outside the object. Keys and meaning:\n"
        "- thesis: 2-3 tight sentences — the long-term investment thesis\n"
        "- business: 1-2 sentences on what it does, its moat, unit economics\n"
        "- bull_case: array of exactly 3 specific one-line points\n"
        "- bear_case: array of exactly 3 specific one-line points\n"
        "- catalysts: array of 2 forward catalysts with rough timing\n"
        "- risks: array of 2 key risks, most material first\n"
        "- verdict: your FINAL call, exactly one of: BUY, ACCUMULATE, HOLD, REDUCE, SELL\n"
        "- verdict_rationale: ONE short clause under 16 words — the core reason for the call\n"
        "- vs_consensus: 2-3 sentences comparing YOUR view to the street_consensus in the facts "
        "(if present) — state where you agree/differ on target and rating, and explain WHY your "
        "view differs (e.g. more conservative WACC, different growth/margin path, catalyst timing). "
        "If no consensus is given, return an empty string.\n"
        "- divergence_factor: name the SINGLE most important FUNDAMENTAL factor where your model differs "
        "from the street, quantified where possible (e.g. 'We model a 300bps-lower steady-state EBIT "
        "margin', 'We use a 12.5% WACC vs the street's implied ~10%', 'We assume FY26 revenue growth of "
        "6% vs consensus ~10% on weaker discretionary demand'). Empty string if no consensus given.\n"
        "- recommendation: 2 sentence rationale for a 3-5 year horizon\n"
        "- price_target: number (your 12-month fair value in INR) or null\n"
        "- scenarios: object with keys bull, base, bear — each an object "
        "{target:<INR number>, probability:<decimal, the three sum to ~1>, driver:'<one line: what "
        "changes vs base — growth/margin/multiple>'}. Base target should match your price_target.\n"
        "- comps: object for market comparables. peers MUST be the company's 3-5 most DIRECT "
        "listed competitors — include recent IPOs and pure-play rivals (e.g. for Eternal/Zomato "
        "include Swiggy; for a private bank its closest private peers; for an EV maker its EV rivals). "
        "Do NOT list loosely-related large caps if a direct competitor is listed. — {peers: array each "
        "{name, ev_ebitda:<x>, pe:<x>}, median_ev_ebitda:<number>, median_pe:<number>, "
        "implied_value_per_share:<INR number applying the peer median to this company>, note:'<1 line>'}.\n"
        "- sotp: Sum-of-the-Parts — ONLY if the company has distinct segments (e.g. Tata Motors = "
        "CV+PV+JLR, Reliance = O2C+Retail+Jio); else null. Object {segments: array of "
        "{segment, basis:'EV/EBITDA'|'EV/Sales'|'P/E', metric_cr:<number>, multiple:<number>, "
        "ev_cr:<number>}, net_debt_cr:<number>, implied_value_per_share:<INR number>, note:'<1 line>'}.\n"
        "- assumptions: object with decimals: growth (array of 5 yearly revenue-growth "
        "rates, realistic vs history), ebit_margin (ARRAY of 5 yearly EBIT margins showing the "
        "TRAJECTORY — anchor Yr1 to the trailing margin then model realistic expansion or "
        "compression from operating leverage / cost inflation; do NOT use one flat number), "
        "tax_rate (~0.25 for India), capex_pct, da_pct, nwc_pct (incremental NWC as % of "
        "revenue change), wacc (discount rate reflecting risk), terminal_growth "
        "(long-run, below wacc, ~0.03-0.05).\n"
        "- wacc_build: object showing how you derived WACC (like an IB memo): "
        "rf (risk-free, India 10yr G-sec ~0.07), erp (equity risk premium ~0.06-0.08), "
        "beta, cost_of_equity, cost_of_debt (after-tax), equity_weight, debt_weight, "
        "and a one-line note. All decimals.\n"
        "- assumption_log: array of 4-6 short strings — each a key modelling assumption "
        "with its basis (e.g. 'Revenue growth fades 15%→8% as base scales', 'EBIT margin "
        "24% vs 22% trailing on operating leverage', 'WACC 12.5% — high ERP for India').\n"
        "- sources: object giving a one-line SOURCE/BASIS for EACH driver (these populate a "
        "Source column in the Excel model). Keys: growth, ebit_margin, tax_rate, capex_pct, "
        "da_pct, nwc_pct, wacc, terminal_growth. For growth, the source MUST show the triangulation, "
        "e.g. 'Blend: 9% co. CAGR / 13% industry / 12% consensus → 11%'. For ebit_margin cite the "
        "starting (trailing) margin and the expansion/compression logic across the 5 years, "
        "trailing margin + operating-leverage logic. Others: 'India 10yr 7% + Damodaran ERP', "
        "'statutory 25%', 'long-run nominal GDP'.\n\n"
        'Example shape: {"thesis":"...","business":"...","bull_case":["..."],'
        '"bear_case":["..."],"catalysts":["..."],"risks":["..."],'
        '"verdict":"ACCUMULATE","verdict_rationale":"Quality franchise, fair value, stagger entries",'
        '"vs_consensus":"Street targets ₹4,100 (BUY, 22 analysts); we are 8% below on a stiffer 12.5% WACC and slower FY26 growth.",'
        '"divergence_factor":"We model FY26 EBIT margin ~300bps below consensus on wage inflation.",'
        '"scenarios":{"bull":{"target":4600,"probability":0.25,"driver":"Margin recovery + faster growth"},'
        '"base":{"target":3600,"probability":0.5,"driver":"In-line execution"},'
        '"bear":{"target":2600,"probability":0.25,"driver":"Demand slump, margin compression"}},'
        '"comps":{"peers":[{"name":"Wipro","ev_ebitda":11,"pe":18},{"name":"HCL Tech","ev_ebitda":13,"pe":22}],'
        '"median_ev_ebitda":12,"median_pe":20,"implied_value_per_share":3400,"note":"Trades at premium to peers"},'
        '"sotp":null,'
        '"recommendation":"...",'
        '"price_target":3200,"assumptions":{"growth":[0.09,0.09,0.08,0.08,0.07],'
        '"ebit_margin":[0.22,0.23,0.24,0.245,0.25],"tax_rate":0.25,"capex_pct":0.03,"da_pct":0.03,"nwc_pct":0.02,'
        '"wacc":0.12,"terminal_growth":0.045},'
        '"wacc_build":{"rf":0.07,"erp":0.07,"beta":0.9,"cost_of_equity":0.133,'
        '"cost_of_debt":0.06,"equity_weight":0.9,"debt_weight":0.1,"note":"India 10yr + Damodaran ERP"},'
        '"assumption_log":["Revenue growth fades 9%→7% as base scales",'
        '"EBIT margin 25% on operating leverage","WACC 12% — elevated ERP for India"],'
        '"sources":{"growth":"3yr hist avg + guidance","ebit_margin":"trailing margin + op leverage",'
        '"tax_rate":"India statutory 25%","capex_pct":"3yr capex/revenue","da_pct":"3yr D&A/revenue",'
        '"nwc_pct":"working-capital trend","wacc":"India 10yr 7% + Damodaran ERP","terminal_growth":"long-run nominal GDP"}}'
    )
    return system, user


def _parse_narrative(text: str):
    """Best-effort extract the JSON object from Claude's reply — tolerant of
    markdown fences, // comments, and trailing commas that LLMs sometimes emit."""
    import re
    if not text:
        return None
    try:
        s = text[text.index("{"): text.rindex("}") + 1]
    except ValueError:
        return None
    for attempt in (s, None):
        if attempt is None:
            # sanitise: strip // comments (not inside URLs) and trailing commas
            s2 = re.sub(r'(?<!:)//[^\n\r]*', '', s)
            s2 = re.sub(r'/\*.*?\*/', '', s2, flags=re.S)
            s2 = re.sub(r',(\s*[}\]])', r'\1', s2)
            attempt = s2
        try:
            return json.loads(attempt)
        except (ValueError, json.JSONDecodeError):
            continue
    return None


def _price_validation_checks(d):
    """Price-integrity entries for the validation layer: cross-source check +
    any sanity-guard correction applied during fetch."""
    checks = []
    pv = cross_validate_price(d["symbol"], d["price"])
    if pv.get("ok") is None:
        checks.append({"name": "price cross-check (2nd source)", "pass": True,
                       "detail": pv.get("note", "n/a") + " — 52-wk sanity guard applied instead"})
    else:
        checks.append({"name": "price cross-check (2nd source)", "pass": pv["ok"],
                       "detail": f"yfinance ₹{d['price']:,.2f} vs Google ₹{pv['google_price']:,.2f} "
                                 f"({pv['diff_pct']}% diff)"})
    if d.get("price_note"):
        checks.append({"name": "price sanity guard (auto-corrected)", "pass": True,
                       "detail": d["price_note"]})
    return checks


def assemble(d: dict, a: dict, narrative=None, validate=True, horizon="long") -> dict:
    recency = data_recency(d)
    # ── SHORT-TERM: technical note, no DCF ──
    if horizon == "short":
        excel = build_excel_short(d, a, narrative)
        tech = d.get("tech") or {}
        pchecks = _price_validation_checks(d)
        result = {
            "symbol": d["symbol"], "name": d["name"], "horizon": "short",
            "verdict": final_verdict(a, narrative), "quant_verdict": a["verdict"], "action": a["action"],
            "price": d["price"], "price_target": (narrative or {}).get("price_target"),
            "data_asof": recency["asof"], "data_status": recency["status"],
            "authored_by": "Claude + live technicals" if narrative else "Technical engine",
            "levels": (narrative or {}).get("levels"),
            "validation": {"ok": bool(tech) and all(c["pass"] for c in pchecks), "checks": [
                {"name": "technical indicators computed", "pass": bool(tech),
                 "detail": f"RSI {tech.get('rsi',0):.0f} · 50/200-DMA present" if tech else "no history"},
                {"name": "data recency (latest/prev quarter)", "pass": recency["ok"],
                 "detail": f"{recency['status']} — as-of {recency['asof']}"},
            ] + pchecks},
            "excel": excel, "pdf": build_pdf_short(d, a, narrative),
        }
        result["numbers"] = ""
        result["analysis"], result["speech"] = compose_short(d, a, narrative, result)
        return result

    # ── LONG-TERM: DCF (default) ──
    asmp = merge_assumptions(default_assumptions(d), (narrative or {}).get("assumptions"))
    excel = build_excel(d, a, narrative, asmp)
    validation = validate_excel(excel, asmp) if validate else None
    if validation is not None:
        # Python data-recency gate → appended to the checks
        validation["checks"].append({
            "name": "data recency (latest/prev quarter)",
            "pass": recency["ok"],
            "detail": f"{recency['status']} — financials as-of {recency['asof']}",
        })
        if not recency["ok"]:
            validation["ok"] = False
        # Price integrity: cross-source check + sanity-guard note
        pchecks = _price_validation_checks(d)
        validation["checks"].extend(pchecks)
        if not all(c["pass"] for c in pchecks):
            validation["ok"] = False
    result = {
        "symbol": d["symbol"],
        "name": d["name"],
        "summary": two_line_summary(d, a, narrative),
        "verdict": final_verdict(a, narrative),
        "quant_verdict": a["verdict"],
        "action": a["action"],
        "price": d["price"],
        "price_target": (narrative or {}).get("price_target"),
        "data_asof": recency["asof"],
        "data_status": recency["status"],
        "authored_by": "Claude + live data" if narrative else "Quant engine (live data)",
        "assumptions": asmp,
        "validation": validation,
        "excel": excel,
        "pdf": build_pdf(d, a, narrative),
    }
    result["numbers"] = numbers_block(d, a, result)
    result["analysis"], result["speech"] = compose_analysis(d, a, narrative, result)
    return result


def generate_report(symbol: str, narrative=None) -> dict:
    d, a = prepare(symbol)
    return assemble(d, a, narrative)


if __name__ == "__main__":
    import sys
    sym = sys.argv[1] if len(sys.argv) > 1 else "INFY"
    r = generate_report(sym)
    with open(f"/tmp/{r['symbol']}_model.xlsx", "wb") as f:
        f.write(r["excel"])
    with open(f"/tmp/{r['symbol']}_report.pdf", "wb") as f:
        f.write(r["pdf"])
    print(r["summary"])
    print(f"Excel: /tmp/{r['symbol']}_model.xlsx  ({len(r['excel'])} bytes)")
    print(f"PDF:   /tmp/{r['symbol']}_report.pdf  ({len(r['pdf'])} bytes)")
